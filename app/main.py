import os
from dotenv import load_dotenv
load_dotenv(override=True)  # override=True 确保加载 .env 文件覆盖 shell 环境变量

# 如果存在 Anthropic/Claude 风格的环境变量，优先映射到 OPENAI_* 以保持向后兼容
_anthropic_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
if _anthropic_key:
    # 仅在未显式设置 OPENAI_* 时进行映射（保留显式 OPENAI_* 优先级）
    os.environ.setdefault("OPENAI_API_KEY", _anthropic_key)
    if os.getenv("ANTHROPIC_BASE_URL", ""):
        os.environ.setdefault("OPENAI_BASE_URL", os.getenv("ANTHROPIC_BASE_URL"))
    if os.getenv("ANTHROPIC_MODEL", ""):
        os.environ.setdefault("OPENAI_MODEL", os.getenv("ANTHROPIC_MODEL"))

from typing import Optional, Dict, Any, AsyncGenerator
from fastapi import FastAPI, Request, Body, UploadFile, File, Form, Depends
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import json
import time
import asyncio
import traceback
from .agent_orchestrator import analyze, replacement_report
from .setup_api import router as setup_router
from sqlalchemy.orm import Session as DBSession
from .database import get_db, init_db
from .auth import get_current_user, get_optional_user
from .models_db import ChatSession as ChatSessionDB, ChatMessage as ChatMessageDB
from .routers.auth_router import router as auth_router
from .routers.admin_router import router as admin_router, _apply_config_to_env
from .session_store import constraint_store

# ── 会话级报告缓存（供 /report/{type} 端点使用）────────────────
# key=session_id, value=(SelectionReport, constraints)
_session_reports: Dict[str, Any] = {}
_session_constraints: Dict[str, Any] = {}
_session_selected_part: Dict[str, Optional[str]] = {}  # session_id → selected part_number
_DEFAULT_SESSION_ID = "__default__"  # 未提供 session_id 时的兜底key
_SERVER_START_TIME = time.time()  # 服务启动时间（用于 /health 上报运行时长）

_SESSION_STORE_MAX = 500  # P0-4: cap in-memory session store to prevent unbounded growth

def _evict_session_stores() -> None:
    """Evict oldest entries when any session dict exceeds the cap."""
    for d in (_session_reports, _session_constraints, _session_selected_part):
        if len(d) > _SESSION_STORE_MAX:
            evict_n = len(d) - _SESSION_STORE_MAX + 50  # evict 50 extra to amortise cost
            for _ in range(evict_n):
                try:
                    d.pop(next(iter(d)))
                except StopIteration:
                    break

# ── 运行时模型管理（支持 API 动态切换，覆盖环境变量）───────────
_runtime_model: Optional[str] = None  # None 表示使用环境变量默认值

AVAILABLE_MODELS = [
    {"id": "deepseek-v4-pro",   "name": "DeepSeek V4 Pro",    "description": "强推理模式，适用于复杂选型分析"},
    {"id": "deepseek-v4-flash", "name": "DeepSeek V4 Flash",  "description": "快速响应模式，适用于标准选型"},
]

def get_active_model() -> str:
    """优先级：运行时 > 环境变量 > DB > 硬编码默认。"""
    if _runtime_model:
        return _runtime_model
    from .llm_config import get_model
    m = get_model()
    if m:
        return m
    try:
        from .database import SessionLocal
        from .models_db import AdminConfig
        _db = SessionLocal()
        cfg = _db.query(AdminConfig).first()
        _db.close()
        if cfg and cfg.llm_model:
            return cfg.llm_model
    except Exception:
        pass
    return "claude-sonnet-5"

app = FastAPI()
app.include_router(setup_router)
app.include_router(auth_router)
app.include_router(admin_router)

@app.on_event("startup")
async def _on_startup():
    init_db()
    # DB 迁移：为旧数据库添加 accumulated_constraints 列（新建时 models_db 已含）
    try:
        from sqlalchemy import text
        from .database import engine
        with engine.connect() as _conn:
            _conn.execute(text("ALTER TABLE chat_sessions ADD COLUMN accumulated_constraints TEXT"))
            _conn.commit()
    except Exception:
        pass  # 列已存在则忽略
    try:
        from .database import SessionLocal
        from .models_db import AdminConfig
        _db = SessionLocal()
        cfg = _db.query(AdminConfig).first()
        if cfg:
            _apply_config_to_env(cfg)
        _db.close()
    except Exception:
        pass

# ── CORS 配置，从环境变量读取白名单（M2）──────────────────────
_cors_origins_raw = os.environ.get(
    "CORS_ORIGINS",
    "http://localhost:3000,http://localhost:3001,https://ezmanbo.online,https://www.ezmanbo.online"
).split(",")
_cors_origins = [o.strip() for o in _cors_origins_raw if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# P2-5: request_id middleware — injects unique ID into every request for end-to-end tracing
import uuid as _uuid
from .log_util import set_request_id as _set_rid

@app.middleware("http")
async def _request_id_middleware(request: Request, call_next):
    rid = request.headers.get("X-Request-ID") or _uuid.uuid4().hex[:8]
    _set_rid(rid)
    response = await call_next(request)
    response.headers["X-Request-ID"] = rid
    return response

class AnalyzeRequest(BaseModel):
    user_input: str
    thinking_depth: str = "default"
    session_id: Optional[str] = None
    pre_constraints: Optional[dict] = None  # 跳过LLM解析，直接使用已提取的结构化约束
    skip_cache: bool = False  # adjustment 意图：跟进语句通用性强，跨会话易误命中缓存

class ReplacementRequest(BaseModel):
    original_part_number: str

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """处理请求验证错误"""
    return JSONResponse(
        status_code=422,
        content={"detail": "请求体格式错误", "errors": exc.errors()},
    )

@app.get("/health")
async def health():
    return {
        "status": "ok",
        "started_at": _SERVER_START_TIME,
        "uptime_s": round(time.time() - _SERVER_START_TIME),
    }

@app.get("/api/models")
async def list_models(current_user=Depends(get_current_user), db: DBSession = Depends(get_db)):
    """返回当前管理员配置的可用模型及激活模型。"""
    from .models_db import AdminConfig
    from .routers.admin_router import PROVIDERS
    cfg = db.query(AdminConfig).first()
    active = (cfg.llm_model if cfg and cfg.llm_model else None) or get_active_model()
    provider_id = cfg.llm_provider if cfg else "custom"
    provider_info = PROVIDERS.get(provider_id, {})
    provider_models = provider_info.get("models", [])
    models = [{"id": m_id, "name": m_id, "description": provider_info.get("name", "")} for m_id in provider_models]
    if not any(m["id"] == active for m in models):
        models.insert(0, {"id": active, "name": active, "description": ""})
    return {"models": models, "active": active, "provider": provider_id}

@app.post("/api/models/switch")
async def switch_model(body: dict = Body(...), current_user=Depends(get_current_user), db: DBSession = Depends(get_db)):
    """切换运行时模型并持久化到 DB。"""
    model_id = body.get("model")
    if not model_id:
        return JSONResponse(status_code=400, content={"detail": "缺少 model 字段"})
    global _runtime_model
    _runtime_model = model_id
    os.environ["ANTHROPIC_MODEL"] = model_id
    os.environ["OPENAI_MODEL"] = model_id
    try:
        from .models_db import AdminConfig
        cfg = db.query(AdminConfig).first()
        if cfg:
            cfg.llm_model = model_id
            db.commit()
    except Exception:
        pass
    return {"status": "ok", "active": model_id}

@app.post("/analyze")
async def analyze_endpoint(
    body: AnalyzeRequest,
    current_user=Depends(get_current_user),
):
    """Pipeline 模式：单次调用，返回完整 SelectionReport JSON。

    ── B4：语义缓存支持 ──
    - 响应头 X-Cache: HIT 表示命中缓存
    - 响应头 X-Cache: MISS 表示未命中缓存
    """
    try:
        session_id = body.session_id or _DEFAULT_SESSION_ID
        # The non-streaming path always executes the pipeline. Raw-text
        # semantic matches are not valid selection-report cache hits.
        cache_header = "MISS"

        # P1: 优先使用 LangGraph 状态机（含 CriticNode + 自动放宽）
        try:
            from .langgraph_orchestrator import run_selection_pipeline
            result = run_selection_pipeline(body.user_input, thinking_depth=body.thinking_depth)
            if result.get("report"):
                report = result["report"]
            else:
                report = analyze(body.user_input, thinking_depth=body.thinking_depth)  # fallback
        except Exception:
            report = analyze(body.user_input, thinking_depth=body.thinking_depth)

        _session_reports[session_id] = report
        _session_constraints[session_id] = report.constraints
        return JSONResponse(
            content=report.dict(),
            headers={"X-Cache": cache_header}
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"detail": f"处理错误: {str(e)}"},
        )

@app.post("/replacement")
async def replacement_endpoint(body: ReplacementRequest, current_user=Depends(get_current_user)):
    """替代器件查找。"""
    try:
        return replacement_report(body.original_part_number)
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"detail": f"处理错误: {str(e)}"},
        )


# ── Agent 端点 ──────────────────────────────────────────────────

class AgentRequest(BaseModel):
    user_input: str
    session_id: Optional[str] = None
    thinking_depth: str = "default"
    accumulated_input: Optional[str] = None   # 多轮澄清时累积的上下文
    has_active_selection: bool = False         # 当前会话是否已有选型结果

@app.post("/agent/chat")
async def agent_chat_endpoint(
    body: AgentRequest,
    current_user=Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    """ReAct Agent 模式：支持多轮对话，返回推理过程 + 工具调用记录。"""
    try:
        from .react_agent import run_agent
        result = run_agent(body.user_input, session_id=body.session_id,
                          thinking_depth=body.thinking_depth)

        # ── DB 会话持久化（游客跳过）────────────────────────
        try:
            if not getattr(current_user, 'is_guest', False):
                sid = body.session_id or _DEFAULT_SESSION_ID
                session_db = db.query(ChatSessionDB).filter(
                    ChatSessionDB.id == sid, ChatSessionDB.user_id == current_user.id
                ).first()
                if not session_db:
                    session_db = ChatSessionDB(id=sid, user_id=current_user.id, title="新的选型")
                    db.add(session_db)
                db.add(ChatMessageDB(session_id=sid, role="user", content=body.user_input[:10000]))
                resp_text = result.get("response", "") if isinstance(result, dict) else ""
                if resp_text:
                    db.add(ChatMessageDB(session_id=sid, role="assistant", content=resp_text[:10000]))
                db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass

        # ── Memory：提取并记录用户称呼 ──────────────────────
        try:
            from .memory import update_user_name
            import re as _re
            name_m = _re.search(r'(?:我是|我叫|叫我)\s*[\u4e00-\u9fff\w]+', body.user_input)
            if name_m:
                name = _re.sub(r'(?:我是|我叫|叫我)\s*', '', name_m.group(0)).strip().rstrip('.,;!。，；！')
                if name and 1 < len(name) <= 10:
                    update_user_name(name)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"agent_chat: 记忆写入失败: {e}")

        return result
    except Exception as e:
        # 如果下游 LLM/服务返回明确的 HTTP 错误（如余额不足 402），把它映射回客户端
        msg = str(e)
        status_code = 500
        try:
            if "Insufficient Balance" in msg or "402" in msg or "invalid_request_error" in msg:
                status_code = 402
        except Exception:
            status_code = 500
        return JSONResponse(
            status_code=status_code,
            content={"detail": f"Agent 错误: {msg}"},
        )

@app.get("/agent/sessions")
async def agent_sessions_endpoint(
    current_user=Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    """列出当前用户的会话（从数据库加载，按用户隔离）。"""
    sessions_db = (
        db.query(ChatSessionDB)
        .filter(ChatSessionDB.user_id == current_user.id)
        .order_by(ChatSessionDB.updated_at.desc())
        .all()
    )
    return {
        "sessions": [
            {"id": s.id, "title": s.title, "message_count": len(s.messages)}
            for s in sessions_db
        ],
        "total": len(sessions_db),
    }


@app.post("/agent/init_session")
async def agent_init_session_endpoint(
    body: dict = Body(...),
    current_user=Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    """创建带预注入上下文的 Agent 会话（无 LLM 调用）。

    直接把选型摘要写入会话历史（AIMessage），
    后续 /agent/chat 调用时 Agent 可基于此上下文作答。

    请求体：
      - session_id: 会话 ID（可选，未提供时自动生成）
      - context: 上下文文本（与 payload 等效）
      - payload: 上下文文本（context 的别名）
      - context_type: 上下文类型标签（如 selection_context, general）
    """
    from .react_agent import get_or_create_session, _sessions
    from langchain_core.messages import AIMessage
    import os as _os

    session_id = body.get("session_id") or _os.urandom(8).hex()
    context = body.get("context") or body.get("payload", "").strip() if isinstance(body, dict) else ""
    ctx_type = body.get("context_type", "general") if isinstance(body, dict) else "general"

    if context:
        history = get_or_create_session(session_id)
        prefix = f"[{ctx_type}]" if ctx_type != "general" else "[选型上下文已同步]"
        history.append(AIMessage(content=f"{prefix}\n{context}"))
        _sessions[session_id] = history

    return {"session_id": session_id, "injected": bool(context), "context_type": ctx_type}


# ── 意图分类端点（三层分流机制）───────────────────────────────

class ClassifyRequest(BaseModel):
    user_input: str
    has_active_selection: bool = False
    accumulated_input: str = ""  # 跨轮累积的约束文本（分号分隔）
    thinking_depth: str = "default"
    session_id: str = _DEFAULT_SESSION_ID

@app.post("/classify")
async def classify_endpoint(
    body: ClassifyRequest,
    current_user=Depends(get_current_user),
):
    """对用户输入进行意图分类，返回富 dict。"""
    from .intent_classifier import classify
    result = classify(
        body.user_input,
        has_active_selection=body.has_active_selection,
        accumulated_input=body.accumulated_input or "",
    )
    # selection_choice：需要对比当前会话的选型报告（服务端状态）
    if result.get("intent") != "selection" and body.has_active_selection:
        import re as _re
        if _re.match(r'^\d{1,2}$', body.user_input.strip()):
            try:
                _report = _session_reports.get(body.session_id)
                if _report:
                    _scored = _report_parts(_report)
                    _idx = int(body.user_input.strip()) - 1
                    if 0 <= _idx < len(_scored):
                        result["intent"] = "selection_choice"
                        result["selected_part"] = _scored[_idx].part.part_number
            except (ValueError, TypeError, AttributeError):
                pass
    return result


# ── 流式对话端点（非选型交互用）──────────────────────────────

@app.post("/agent/chat/stream")
async def agent_chat_stream_endpoint(
    body: AgentRequest,
    current_user=Depends(get_current_user),
):
    """轻量流式对话 — 不触发选型流水线，仅 ReAct Agent 自然语言交互。"""

    async def _stream_chat() -> AsyncGenerator[str, None]:
        import time as _time
        t0 = _time.time()
        try:
            from .react_agent import run_agent
            from .intent_classifier import classify
            from .llm_client import call_openai_chat
            from .llm_config import get_model
            import json as _json

            yield f"event: start\ndata: {_json.dumps({'status': 'agent_thinking'})}\n\n"

            # ── 推理过程展示：预调用 LLM 获取 reasoning_content ────
            if body.thinking_depth != "off":
                try:
                    model = get_model() or "claude-sonnet-5"
                    reasoning_resp = call_openai_chat(
                        messages=[{"role": "user", "content": body.user_input}],
                        model=model,
                        thinking_depth=body.thinking_depth,
                    )
                    rc = reasoning_resp.get("reasoning_content")
                    if rc:
                        # 分段发射，每段约 200 字符，让前端逐步显示
                        chunk_size = 200
                        for i in range(0, len(rc), chunk_size):
                            chunk = rc[i:i + chunk_size]
                            yield f"event: thinking_delta\ndata: {_json.dumps({'stage': 'deepseek', 'text': chunk}, ensure_ascii=False)}\n\n"
                            await asyncio.sleep(0.02)  # 避免前端一次性接收太多
                except Exception:
                    pass  # 推理展示失败不影响主流程

            result = run_agent(body.user_input, session_id=body.session_id,
                              thinking_depth=body.thinking_depth)

            # ── 思考流：暴露 ReAct 工具调用链 ────────────────────
            if body.thinking_depth != "off":
                tool_calls = result.get("tool_calls", [])
                for tc in tool_calls:
                    tool_name = tc.get("tool", "unknown")
                    tool_args = tc.get("args", {})
                    args_summary = ", ".join(
                        f"{k}={str(v)[:60]}" for k, v in tool_args.items()
                    )
                    yield f"event: thinking_delta\ndata: {_json.dumps({'stage': 'agent', 'text': f'调用工具：{tool_name}({args_summary})'}, ensure_ascii=False)}\n\n"
                    raw_result = tc.get("result", "")
                    if raw_result and body.thinking_depth in ("contemplation", "exhaustive"):
                        preview = str(raw_result)[:120].replace("\n", " ")
                        yield f"event: thinking_delta\ndata: {_json.dumps({'stage': 'agent', 'text': f'工具返回：{preview}'}, ensure_ascii=False)}\n\n"

            # 提取文本回复
            text = ""
            if isinstance(result, dict):
                text = result.get("response") or result.get("output") or str(result)
            else:
                text = str(result)

            yield f"event: text_delta\ndata: {_json.dumps({'text': text})}\n\n"

            # ── Memory：提取并记录用户称呼 ────────────────────
            try:
                from .memory import update_user_name, get_user_context
                import re as _re
                name_m = _re.search(r'(?:我是|我叫|叫我|name is|I am|I\'m)\s*[\x80-\uffff\w]+', body.user_input)
                if name_m:
                    raw = name_m.group(0)
                    name = _re.sub(r'(?:我是|我叫|叫我|name is|I am|I\'m)\s*', '', raw).strip().rstrip('.,;!。，；！')
                    if name and len(name) <= 10:
                        update_user_name(name)
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"agent_chat_stream: 记忆写入失败: {e}")

            elapsed = round(_time.time() - t0, 2)
            yield f"event: done\ndata: {_json.dumps({'elapsed_s': elapsed, 'intent': 'chat'})}\n\n"
        except Exception as e:
            # 记录并按情况将下游错误映射到结构化 SSE 事件
            try:
                from .log_util import log_error
                log_error("main.agent_chat_stream", e, "streaming_agent")
            except Exception:
                pass
            msg = str(e)
            code = None
            # 根据错误类型返回用户可读的中文提示
            if "Insufficient Balance" in msg or "402" in msg:
                code = 402
                msg = "API 余额不足，请联系管理员充值"
            elif "invalid_request_error" in msg or "400" in msg:
                code = 400
                msg = "模型请求参数错误，请稍后重试"
            elif "Timeout" in msg or "timed out" in msg.lower():
                code = 408
                msg = "模型响应超时，请简化问题后重试"
            elif "429" in msg or "Rate limit" in msg or "rate_limit" in msg:
                code = 429
                msg = "请求过于频繁，请稍候再试"
            elif "401" in msg or "Unauthorized" in msg:
                code = 401
                msg = "API 密钥无效，请联系管理员检查配置"
            elif "404" in msg:
                code = 404
                msg = "模型服务暂不可用，请稍后重试"
            elif "ConnectionError" in msg or "connection" in msg.lower():
                code = 503
                msg = "网络连接异常，请检查网络后重试"
            payload = {"message": msg}
            if code:
                payload["code"] = code
            yield f"event: error\ndata: {_json.dumps(payload)}\n\n"

    return StreamingResponse(
        _stream_chat(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── 统一流式入口（合并意图分类 + 路由）──────────────────────────────

async def _with_heartbeat(gen: AsyncGenerator[str, None], interval: float = 15.0) -> AsyncGenerator[str, None]:
    """包装 SSE 生成器，每 interval 秒发送一次 keep-alive 注释，防止浏览器/Nginx 超时断开。"""
    queue: asyncio.Queue = asyncio.Queue(maxsize=64)
    _sentinel = object()

    async def _producer():
        try:
            async for item in gen:
                await queue.put(item)
        finally:
            await queue.put(_sentinel)

    task = asyncio.create_task(_producer())
    try:
        while True:
            try:
                item = await asyncio.wait_for(queue.get(), timeout=interval)
                if item is _sentinel:
                    return
                yield item
            except asyncio.TimeoutError:
                yield ": ping\n\n"
    finally:
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass


async def _stream_agent_chat(body: AgentRequest) -> AsyncGenerator[str, None]:
    """模块级 Agent 对话流式生成器（供 _stream_unified 复用）。"""
    import time as _time, json as _json
    t0 = _time.time()
    try:
        from .react_agent import run_agent

        yield f"event: start\ndata: {_json.dumps({'status': 'agent_thinking'})}\n\n"

        # 注入已累积的选型参数作为上下文（避免 LLM 重复追问已给出的参数）
        sid = body.session_id or _DEFAULT_SESSION_ID
        accumulated_c = constraint_store.get(sid)
        context_note = ""
        if accumulated_c:
            _t_map = {"buck": "降压", "boost": "升压", "ldo": "线性稳压(LDO)", "buck_boost": "升降压"}
            param_parts = []
            if accumulated_c.get("topology"):
                param_parts.append(f"电路拓扑：{_t_map.get(accumulated_c['topology'], accumulated_c['topology'])}")
            if accumulated_c.get("input_voltage_nominal_v"):
                param_parts.append(f"输入电压：{accumulated_c['input_voltage_nominal_v']}V")
            if accumulated_c.get("output_voltage_v"):
                param_parts.append(f"输出电压：{accumulated_c['output_voltage_v']}V")
            if accumulated_c.get("output_current_a"):
                iout = accumulated_c["output_current_a"]
                param_parts.append(f"输出电流：{int(iout*1000)}mA" if iout < 1 else f"输出电流：{iout}A")
            if accumulated_c.get("grade"):
                g = {"automotive": "车规(AEC-Q100)", "industrial": "工业级", "commercial": "商业级"}
                param_parts.append(f"等级：{g.get(accumulated_c['grade'], accumulated_c['grade'])}")
            if param_parts:
                context_note = (
                    "[会话已收集的选型参数]\n"
                    + "\n".join(f"- {p}" for p in param_parts)
                    + "\n以上参数已记录，不要再次追问，只收集缺少的参数。"
                )

        # Run synchronous agent in thread pool to avoid blocking the event loop
        result = await asyncio.to_thread(
            run_agent, body.user_input,
            session_id=body.session_id,
            thinking_depth=body.thinking_depth,
            context_note=context_note,
        )

        # Stream thinking content returned from the single LLM call
        if body.thinking_depth != "off":
            rc = result.get("reasoning_content") if isinstance(result, dict) else None
            if rc:
                for i in range(0, len(rc), 200):
                    yield f"event: thinking_delta\ndata: {_json.dumps({'stage': 'agent', 'text': rc[i:i+200]}, ensure_ascii=False)}\n\n"
                    await asyncio.sleep(0.01)
            for tc in (result.get("tool_calls") or []):
                tool_name = tc.get("tool", "unknown")
                args_summary = ", ".join(f"{k}={str(v)[:60]}" for k, v in tc.get("args", {}).items())
                yield f"event: thinking_delta\ndata: {_json.dumps({'stage': 'agent', 'text': f'调用工具：{tool_name}({args_summary})'}, ensure_ascii=False)}\n\n"

        text = ""
        if isinstance(result, dict):
            text = result.get("response") or result.get("output") or str(result)
        else:
            text = str(result)

        yield f"event: text_delta\ndata: {_json.dumps({'text': text})}\n\n"

        try:
            from .memory import update_user_name
            import re as _re
            name_m = _re.search(r'(?:我是|我叫|叫我|name is|I am|I\'m)\s*[\x80-￿\w]+', body.user_input)
            if name_m:
                raw = name_m.group(0)
                name = _re.sub(r'(?:我是|我叫|叫我|name is|I am|I\'m)\s*', '', raw).strip().rstrip('.,;!。，；！')
                if name and len(name) <= 10:
                    update_user_name(name)
        except Exception:
            pass

        yield f"event: done\ndata: {_json.dumps({'elapsed_s': round(_time.time() - t0, 2), 'intent': 'chat'})}\n\n"

    except Exception as e:
        try:
            from .log_util import log_error
            log_error("stream_agent_chat", e, "streaming_agent")
        except Exception:
            pass
        msg = str(e)
        code = None
        if "Insufficient Balance" in msg or "402" in msg:
            code, msg = 402, "API 余额不足，请联系管理员充值"
        elif "Timeout" in msg or "timed out" in msg.lower():
            code, msg = 408, "模型响应超时，请简化问题后重试"
        elif "429" in msg or "rate_limit" in msg:
            code, msg = 429, "请求过于频繁，请稍候再试"
        elif "401" in msg or "Unauthorized" in msg:
            code, msg = 401, "API 密钥无效，请联系管理员检查配置"
        elif "ConnectionError" in msg or "connection" in msg.lower():
            code, msg = 503, "网络连接异常，请检查网络后重试"
        payload = {"message": msg}
        if code:
            payload["code"] = code
        yield f"event: error\ndata: {json.dumps(payload)}\n\n"


def _build_constraints_text(constraints: dict, original_input: str) -> str:
    """将累积的结构化约束序列化为自然语言文本，供需求解析器使用。"""
    parts: list = []
    t_map = {"buck": "Buck降压", "boost": "Boost升压", "ldo": "LDO线性稳压",
             "buck_boost": "Buck-Boost"}
    if constraints.get("topology"):
        parts.append(t_map.get(constraints["topology"], constraints["topology"]))
    vin = constraints.get("input_voltage_nominal_v")
    if vin is None and constraints.get("input_voltage_max_v"):
        vin = constraints["input_voltage_max_v"]
    if vin:
        if constraints.get("input_voltage_min_v") and constraints["input_voltage_min_v"] != vin:
            parts.append(f"输入电压{constraints['input_voltage_min_v']}V~{vin}V")
        else:
            parts.append(f"输入电压{vin}V")
    if constraints.get("output_voltage_v"):
        parts.append(f"输出电压{constraints['output_voltage_v']}V")
    if constraints.get("output_current_a"):
        iout = constraints["output_current_a"]
        parts.append(f"输出电流{int(iout * 1000)}mA" if iout < 1 else f"输出电流{iout}A")
    if constraints.get("temperature_min_c") is not None and constraints.get("temperature_max_c") is not None:
        parts.append(f"温度范围{constraints['temperature_min_c']}~{constraints['temperature_max_c']}°C")
    g_map = {"automotive": "车规级AEC-Q100", "industrial": "工业级", "commercial": "商业级"}
    if constraints.get("grade"):
        parts.append(g_map.get(constraints["grade"], constraints["grade"]))
    if constraints.get("package_preference"):
        parts.append(f"封装{constraints['package_preference']}")
    if parts:
        return "，".join(parts) + "。" + original_input
    return original_input


async def _stream_unified(body: AgentRequest, dual_model_enabled: bool = False) -> AsyncGenerator[str, None]:
    """统一流式入口：在流内部完成意图分类，再路由到相应处理链路。

    事件顺序：
      start → intent → (selection: cache_hit / stage / parse_done / ... / done)
                      (chat: thinking_delta / text_delta / done)
                      (selection_choice / clarify: text_delta / done)
      error（任意阶段异常时）
    """
    import time as _t, json as _j
    t0 = _t.time()
    try:
        yield _yield_sse("start", {"session_id": body.session_id or ""})

        # ── 服务端约束累积（解决多轮参数逐步提供时丢失问题）─────────
        from .constraint_checker import extract_constraints, merge_constraints, check_completeness
        from .intent_classifier import _is_fast_chat, _is_fast_adjustment, classify, extract_adjustment
        import re as _re

        sid = body.session_id or _DEFAULT_SESSION_ID

        # 从 DB 恢复（服务器重启后内存清零时使用）
        if not constraint_store.contains(sid) and body.session_id:
            try:
                import json as _jc
                from .database import SessionLocal as _SLc
                from .models_db import ChatSession as _CSc
                _dbc = _SLc()
                try:
                    _cs = _dbc.query(_CSc).filter(_CSc.id == sid).first()
                    if _cs and getattr(_cs, 'accumulated_constraints', None):
                        constraint_store.set(sid, _jc.loads(_cs.accumulated_constraints))
                finally:
                    _dbc.close()
            except Exception:
                pass

        new_c = extract_constraints(body.user_input)
        accumulated_c = constraint_store.get(sid)
        merged_c = merge_constraints(accumulated_c, new_c)
        # Always persist merged state — even if no new params extracted this turn
        constraint_store.set(sid, merged_c)
        # 持久化到 DB（重启后不丢失）
        if body.session_id:
            try:
                import json as _jp
                from .database import SessionLocal as _SLp
                from .models_db import ChatSession as _CSp
                _dbp = _SLp()
                try:
                    _csp = _dbp.query(_CSp).filter(_CSp.id == sid).first()
                    if _csp:
                        _csp.accumulated_constraints = _jp.dumps(merged_c, ensure_ascii=False)
                        _dbp.commit()
                finally:
                    _dbp.close()
            except Exception:
                pass

        # ── 意图路由（约束完整性优先，减少独立 LLM 分类调用）────────
        cls: dict = {"intent": "chat", "merged_input": body.user_input}

        _ORDINAL_MAP = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8}

        def _extract_ordinal(text: str):
            """从自然语言中提取序号，如"第三个方案"→3，"选2"→2"""
            m = _re.search(
                r'(?:选择?|确认|要|用|选用|第)\s*([一二三四五六七八\d])\s*(?:个|号|款|方案|个方案)?',
                text
            )
            if m:
                raw = m.group(1).strip()
                return _ORDINAL_MAP.get(raw) or (int(raw) if raw.isdigit() else None)
            # bare ordinal "第三个"
            m2 = _re.match(r'^\s*第\s*([一二三四五六七八\d])\s*(?:个|号|款|方案)?', text.strip())
            if m2:
                raw = m2.group(1).strip()
                return _ORDINAL_MAP.get(raw) or (int(raw) if raw.isdigit() else None)
            return None

        # 1. 数字或自然语言选择（优先级最高）
        _sel_idx = None
        if body.has_active_selection:
            if _re.match(r'^\d{1,2}$', body.user_input.strip()):
                _sel_idx = int(body.user_input.strip())
            else:
                _sel_idx = _extract_ordinal(body.user_input)

        if _sel_idx is not None:
            try:
                _report = _session_reports.get(sid)
                if _report:
                    _scored = _report_parts(_report)
                    _idx = _sel_idx - 1
                    if 0 <= _idx < len(_scored):
                        cls = {"intent": "selection_choice",
                               "selected_part": _scored[_idx].part.part_number}
            except (ValueError, TypeError, AttributeError):
                pass

        # 2. 出域快速拦截（在 fast_chat 之前，避免被快速路径绕过）
        if cls["intent"] == "chat" and _sel_idx is None:
            from .intent_classifier import _is_out_of_scope, _OUT_OF_SCOPE_REPLY
            if _is_out_of_scope(body.user_input):
                cls = {"intent": "out_of_scope", "clarify_response": _OUT_OF_SCOPE_REPLY}

        # 2b. 身份问题快速拦截（直接返回固定回复，防止底层模型泄露身份）
        _IDENTITY_PATTERNS = ("你是谁", "你是什么", "你叫什么", "你是claude", "你是ai",
                               "who are you", "what are you", "你是gpt", "你是哪家",
                               "谁开发的", "你是什么ai", "介绍一下你自己", "自我介绍")
        _t_lower = body.user_input.strip().lower()
        if cls["intent"] == "chat" and any(p in _t_lower for p in _IDENTITY_PATTERNS):
            yield _yield_sse("intent", {"intent": "chat", "confidence": 1.0})
            yield _yield_sse("text_delta", {"text": "我是 eZmanbo，智能电子元器件选型助理，专注于电源管理 IC、DC-DC 转换器、LDO 等元器件的选型与评估。"})
            yield _yield_sse("done", {"elapsed_s": 0.0, "intent": "chat"})
            return

        # 3. 纯问候/对话快速路径（无需参数）
        if cls["intent"] == "chat" and _is_fast_chat(body.user_input) and _sel_idx is None:
            cls = {"intent": "chat"}

        # 3. 调整指令（有活跃选型结果时）
        elif cls["intent"] == "chat" and body.has_active_selection and _is_fast_adjustment(body.user_input):
            cls = {"intent": "adjustment",
                   "adjustments": extract_adjustment(body.user_input)}

        # 4. 基于累积约束完整性判断（主路径：消除冗余 LLM 分类调用）
        elif cls["intent"] == "chat" and merged_c:
            is_complete, missing_p0, missing_p1 = check_completeness(merged_c)
            if is_complete:
                cls = {"intent": "selection", "merged_input": body.user_input, "missing_p1": missing_p1}
            else:
                cls = {"intent": "clarify", "missing_p0": missing_p0}

        # 5. 兜底：LLM 分类（处理纯描述性问题、模糊意图等）
        else:
            cls = classify(
                body.user_input,
                has_active_selection=body.has_active_selection,
                accumulated_input=body.accumulated_input or "",
                session_id=sid,
            )

        intent: str = cls.get("intent", "chat")

        yield _yield_sse("intent", {
            "intent": intent,
            "selected_part": cls.get("selected_part"),
            "merged_input": cls.get("merged_input"),
            "clarify_response": cls.get("clarify_response"),
            "adjustments": cls.get("adjustments"),
            "no_spec_params": cls.get("no_spec_params"),
            "confidence": cls.get("confidence", 1.0),
        })

        # ── 数字选择：器件确认 ─────────────────────────────────────
        if intent == "selection_choice" and cls.get("selected_part"):
            pn = cls["selected_part"]
            sid = body.session_id or _DEFAULT_SESSION_ID
            _session_selected_part[sid] = pn
            yield _yield_sse("done", {
                "elapsed_s": round(_t.time() - t0, 2),
                "intent": "selection_choice",
                "selected_part": pn,
            })
            return

        # ── 参数澄清：LLM 生成自然回复，显示已收集的参数避免重复追问 ──
        if intent == "clarify":
            from .constraint_checker import build_clarification_response
            clarify_text, updated_c = await asyncio.to_thread(
                build_clarification_response, body.user_input, merged_c  # pass merged_c, not just accumulated_c
            )
            # Merge updated_c INTO merged_c to avoid overwriting already-collected params
            if updated_c:
                final_c = merge_constraints(merged_c, updated_c)
                constraint_store.set(sid, final_c)
            missing_p0_fields = list(cls.get("missing_p0") or [])
            accumulated_for_sse = merge_constraints(merged_c, updated_c) if updated_c else merged_c
            yield _yield_sse("clarify_fields", {"missing_p0": missing_p0_fields, "accumulated": accumulated_for_sse or {}})
            yield _yield_sse("text_delta", {"text": clarify_text})
            yield _yield_sse("done", {"elapsed_s": round(_t.time() - t0, 2), "intent": "clarify"})
            return

        # ── 选型 / 调整：完整分析流水线 ────────────────────────────
        if intent in ("selection", "adjustment"):
            # 优先使用服务端累积的结构化约束重建完整分析文本
            if merged_c and intent == "selection":
                merged = _build_constraints_text(merged_c, body.user_input)
                constraint_store.pop(sid)  # 选型触发后清空累积
            else:
                merged = cls.get("merged_input") or (
                    f"{body.accumulated_input}; {body.user_input}"
                    if body.accumulated_input else body.user_input
                )
            analyze_body = AnalyzeRequest(
                user_input=merged,
                thinking_depth=body.thinking_depth,
                session_id=body.session_id,
                pre_constraints=merged_c if intent == "selection" and merged_c else None,
                skip_cache=(intent == "adjustment"),
            )
            async for ev in _stream_analyze(analyze_body, dual_model_enabled=dual_model_enabled):
                yield ev

            # P1-2: 选型完成后，如果有缺失的 P1 字段，追问以优化后续结果
            if intent == "selection" and cls.get("missing_p1"):
                from .constraint_checker import generate_clarification_questions
                p1_questions = generate_clarification_questions([], cls["missing_p1"])
                if p1_questions:
                    yield _yield_sse("p1_followup", {
                        "questions": p1_questions,
                        "hint": "补充以下信息可进一步优化选型结果："
                    })
            return

        # ── 出域请求：直接返回固定拒绝文本，不进 Agent ────────────
        if intent == "out_of_scope":
            reply = cls.get("clarify_response") or "抱歉，我只能协助电子元器件选型相关问题。如需选型，请描述您的电压/电流/拓扑需求。"
            yield _yield_sse("text_delta", {"text": reply})
            yield _yield_sse("done", {"elapsed_s": round(_t.time() - t0, 2), "intent": "out_of_scope"})
            return

        # ── 其他（chat / replacement / general）: Agent 对话 ──────
        async for ev in _stream_agent_chat(body):
            yield ev

    except Exception as e:
        try:
            from .log_util import log_error
            log_error("stream_unified", e, "streaming_unified")
        except Exception:
            pass
        yield _yield_sse("error", {"message": str(e)})


@app.post("/chat/stream")
async def chat_stream_endpoint(
    body: AgentRequest,
    current_user=Depends(get_current_user),
    db: DBSession = Depends(get_db),
):
    """统一流式对话端点：意图分类 + 路由均在服务端完成，前端单连接处理所有场景。"""
    async def _gen():
        async for ev in _stream_unified(body, dual_model_enabled=bool(getattr(current_user, 'dual_model_enabled', False))):
            yield ev
        # DB 持久化（游客跳过）
        if not getattr(current_user, "is_guest", False):
            try:
                sid = body.session_id or _DEFAULT_SESSION_ID
                session_db = db.query(ChatSessionDB).filter(
                    ChatSessionDB.id == sid, ChatSessionDB.user_id == current_user.id
                ).first()
                if not session_db:
                    session_db = ChatSessionDB(id=sid, user_id=current_user.id, title="新的对话")
                    db.add(session_db)
                db.add(ChatMessageDB(session_id=sid, role="user", content=body.user_input[:10000]))
                db.commit()
                # P0-3: first message — session just created, persist any already-accumulated constraints
                if constraint_store.contains(sid) and not getattr(session_db, "accumulated_constraints", None):
                    import json as _jfix
                    session_db.accumulated_constraints = _jfix.dumps(
                        constraint_store.get(sid), ensure_ascii=False
                    )
                    db.commit()
            except Exception:
                try:
                    db.rollback()
                except Exception:
                    pass

    return StreamingResponse(
        _with_heartbeat(_gen()),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )

def _build_interaction_prompt(report) -> str:
    """选型完成后生成交互引导文案。"""
    recs = report.recommended_parts or []
    if not recs:
        return "\n\n> ⚠ 未找到满足条件的推荐器件，建议放宽约束后重试。"

    top1 = recs[0]
    pn = top1.part.part_number if hasattr(top1.part, "part_number") else "?"
    mfr = top1.part.manufacturer if hasattr(top1.part, "manufacturer") else ""
    score = int(top1.score.total_score) if hasattr(top1.score, "total_score") else 0
    risk = report.risks.overall_risk_level if hasattr(report.risks, "overall_risk_level") else "?"
    total_count = len(recs) + len([
        s for s in (report.candidates or [])
        if getattr(s, "recommendation_level", "") == "backup"
    ])

    return f"""
---

> 共筛选出 **{total_count}** 个候选器件，首选推荐 **#{1} {pn}**（{mfr}），综合评分 **{score}** 分。

请回复编号（如 `1`）选定器件，或告诉我需要调整哪些参数。"""


def _rebuild_summary_from_cached_dict(cached_report: dict) -> str:
    """从缓存结果生成简洁的自然语言摘要（不使用模板表格）。"""
    # candidates 优先，兼容旧缓存中 scored_parts 字段名
    candidates = (cached_report.get("candidates") or
                  cached_report.get("scored_parts") or [])
    recommended = [c for c in candidates
                   if (c.get("recommendation_level") if isinstance(c, dict) else None) == "recommended"]
    if recommended:
        r0 = recommended[0]
        p = r0.get("part", {}) if isinstance(r0, dict) else {}
        s = r0.get("score", {}) if isinstance(r0, dict) else {}
        pn = p.get("part_number", "?") if isinstance(p, dict) else "?"
        mfr = p.get("manufacturer", "") if isinstance(p, dict) else ""
        score = int(s.get("total_score", 0)) if isinstance(s, dict) else 0
        return (
            f"（缓存结果）共 {len(candidates)} 款候选器件，"
            f"首选推荐 **{pn}**（{mfr}，综合评分 {score} 分）。"
            "详细评分请在右侧面板查看，也可告诉我选第1款或换国产替代。"
        )
    return f"（缓存结果）共 {len(candidates)} 款候选器件，详细评分请在右侧面板查看。"


def _safe_serialize(obj):
    """安全序列化 Pydantic/dataclass 对象为 JSON 兼容 dict。"""
    if hasattr(obj, 'dict'):
        return _safe_serialize(obj.dict())
    if hasattr(obj, 'model_dump'):
        return _safe_serialize(obj.model_dump())
    if isinstance(obj, dict):
        return {k: _safe_serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_safe_serialize(v) for v in obj]
    return obj


def _yield_sse(event: str, data: dict) -> str:
    """构造 SSE 事件字符串（模块级，供各阶段子函数复用）。"""
    return f"event: {event}\ndata: {json.dumps(_safe_serialize(data), ensure_ascii=False)}\n\n"


def _report_parts(report):
    """Return the current report's complete candidate collection.

    SelectionReport uses ``candidates`` as its canonical field. The fallback
    keeps old persisted/cache payloads readable without reviving ``scored_parts``
    as a public schema field.
    """
    if isinstance(report, dict):
        return report.get("candidates") or report.get("recommended_parts") or report.get("scored_parts") or []
    return (getattr(report, "candidates", None)
            or getattr(report, "recommended_parts", None)
            or getattr(report, "scored_parts", None)
            or [])


def _part_number(scored_part):
    if isinstance(scored_part, dict):
        part = scored_part.get("part") or scored_part
        return part.get("part_number", "") if isinstance(part, dict) else ""
    return getattr(getattr(scored_part, "part", None), "part_number", "")


# ═══════════════════════════════════════════════════════════════════
# 阶段子函数（拆分自 _stream_analyze，降低圈复杂度）
# 每个函数返回 (结果, sse_events_list)
# ═══════════════════════════════════════════════════════════════════

async def _stream_stage1_parse(req: AnalyzeRequest, loop) -> tuple:
    """Stage 1：需求解析 → (requirement, events)

    快速路径：若 req.pre_constraints 已由上游填充，直接构建 RequirementConstraints 跳过 LLM 调用。
    """
    events: list = []
    events.append(_yield_sse("stage", {"stage": "parse", "status": "started"}))
    from .requirement_parser import parse_requirement

    if req.pre_constraints:
        # 快速路径：多轮累积的结构化约束已完整，无需额外 LLM 解析
        from .schemas import RequirementConstraints as _RC
        try:
            requirement = _RC(raw_input=req.user_input, **{
                k: v for k, v in req.pre_constraints.items() if v is not None
            })
        except Exception:
            requirement = await loop.run_in_executor(None, parse_requirement, req.user_input)
    else:
        requirement = await loop.run_in_executor(None, parse_requirement, req.user_input)
    events.append(_yield_sse("parse_done", {
        "status": "需求解析完成",
        "constraints": _safe_serialize(requirement),
        "fields_parsed": sum(1 for v in [
            requirement.category, requirement.topology,
            requirement.input_voltage_nominal_v, requirement.output_voltage_v,
            requirement.output_current_a, requirement.temperature_min_c,
            requirement.temperature_max_c, requirement.grade,
        ] if v is not None),
    }))
    if req.thinking_depth != "off":
        think_parts = []
        if requirement.topology:
            think_parts.append(f"拓扑：{requirement.topology}")
        if requirement.input_voltage_nominal_v:
            think_parts.append(f"Vin={requirement.input_voltage_nominal_v}V")
        if requirement.output_voltage_v:
            think_parts.append(f"Vout={requirement.output_voltage_v}V")
        if requirement.output_current_a:
            think_parts.append(f"Iout={requirement.output_current_a}A")
        if requirement.temperature_min_c is not None and requirement.temperature_max_c is not None:
            think_parts.append(f"温度范围 {requirement.temperature_min_c}~{requirement.temperature_max_c}°C")
        if requirement.grade:
            think_parts.append(f"等级：{requirement.grade}")
        events.append(_yield_sse("thinking_delta", {
            "stage": "parse",
            "text": "约束提取：" + ("、".join(think_parts) if think_parts else "参数不完整，将进行宽松搜索"),
        }))
    return requirement, events


async def _stream_stage2_search(req: AnalyzeRequest, requirement, loop) -> tuple:
    """Stage 2：器件搜索 + 详情富化 + 候选放宽 → (candidates, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "search", "status": "started"}))
    from .multi_agent import parallel_search_parts, parallel_enrich_candidates
    events.append(_yield_sse("agent_activity", {
        "agent": "SearchOrchestrator", "status": "启动并行搜索", "phase": "search",
        "detail": "多 SearchAgent 并发查询 eZ-PLM",
    }))
    candidates = await parallel_search_parts(requirement)
    events.append(_yield_sse("search_done", {
        "status": "搜索完成",
        "candidate_count": len(candidates),
        "sources": list(set(getattr(c, "source", "unknown") for c in candidates)),
    }))
    if candidates:
        events.append(_yield_sse("agent_activity", {
            "agent": "EnrichOrchestrator", "status": "并行富化器件详情", "phase": "enrich",
        }))
        candidates = await parallel_enrich_candidates(candidates, max_enrich=6)
    if req.thinking_depth != "off":
        if candidates:
            mfrs = list(dict.fromkeys(
                getattr(c, "manufacturer", "") for c in candidates
                if getattr(c, "manufacturer", "")
            ))[:3]
            mfr_str = "、".join(mfrs) if mfrs else "多品牌"
            events.append(_yield_sse("thinking_delta", {
                "stage": "search",
                "text": f"检索到 {len(candidates)} 个候选，来自 {mfr_str} 等厂商，按参数匹配→供应链→成本→国产化率四维评分",
            }))
        else:
            events.append(_yield_sse("thinking_delta", {
                "stage": "search",
                "text": "未找到精确匹配器件，将放宽电流约束后重新搜索",
            }))
    if not candidates:
        events.append(_yield_sse("warning", {"message": "未找到匹配器件，请检查需求或扩充数据源"}))
    if len(candidates) < 3 and not getattr(requirement, 'output_current_a', None) is None:
        old_iout = requirement.output_current_a
        requirement.output_current_a = round(old_iout * 0.8, 2)
        events.append(_yield_sse("stage", {
            "stage": "search",
            "status": f"candidates<3, relaxing Iout {old_iout}A->{requirement.output_current_a}A",
        }))
        candidates = await parallel_search_parts(requirement)
        requirement.output_current_a = old_iout
    return candidates, events


async def _stream_stage3_score(req: AnalyzeRequest, requirement, candidates, loop) -> tuple:
    """Stage 3：评分计算 → (scored, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "score", "status": "started", "total": len(candidates)}))
    from .scoring import score_candidates
    scored = await loop.run_in_executor(None, score_candidates, requirement, candidates)
    for i, s in enumerate(scored):
        events.append(_yield_sse("score_update", {
            "status": f"评分完成: {s.part.part_number}",
            "index": i + 1, "total": len(scored),
            "part_number": s.part.part_number,
            "manufacturer": s.part.manufacturer,
            "total_score": s.score.total_score,
            "parameter_match_score": s.score.parameter_match_score,
            "recommendation_level": s.recommendation_level,
            "scoring_mode": s.score.scoring_mode,
        }))
    if req.thinking_depth != "off" and scored:
        top = scored[0]
        score_detail = (
            f"参数 {int(top.score.parameter_match_score)}"
            f"+供应 {int(top.score.supply_risk_score)}"
            f"+成本 {int(top.score.cost_score)}"
            f"+国产 {int(top.score.domestic_score)}"
            f"= {int(top.score.total_score)} 分"
        )
        events.append(_yield_sse("thinking_delta", {
            "stage": "score",
            "text": f"首选：{top.part.part_number}（{top.part.manufacturer or '—'}）{score_detail}",
        }))
        if req.thinking_depth in ("contemplation", "exhaustive") and top.score.reasons:
            reasons_str = "；".join(top.score.reasons[:3])
            events.append(_yield_sse("thinking_delta", {
                "stage": "score",
                "text": f"评分依据：{reasons_str}",
            }))
    return scored, events


async def _stream_stage4_evidence(req: AnalyzeRequest, scored, requirement, loop) -> tuple:
    """Stage 4：证据构建 → (evidence, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "evidence", "status": "started"}))
    from .evidence import build_evidence
    evidence = await loop.run_in_executor(None, build_evidence, scored, requirement)
    avg_conf = round(sum(e.confidence for e in evidence) / len(evidence), 3) if evidence else 0.0
    evidence_items = []
    for e in evidence:
        evidence_items.append({
            "part_number": e.part_number,
            "claim": e.claim,
            "evidence_type": e.evidence_type,
            "source_field": e.source_field,
            "confidence": e.confidence,
            "need_human_review": getattr(e, "need_human_review", False),
        })
    events.append(_yield_sse("evidence_done", {
        "status": "证据链构建完成",
        "evidence_count": len(evidence),
        "avg_confidence": avg_conf,
        "evidence_items": evidence_items,
    }))
    if req.thinking_depth != "off" and evidence:
        high_conf = sum(1 for e in evidence if e.confidence >= 0.8)
        need_review = sum(1 for e in evidence if getattr(e, "need_human_review", False))
        events.append(_yield_sse("thinking_delta", {
            "stage": "evidence",
            "text": (
                f"证据链：{len(evidence)} 条参数符合性记录"
                f"，高置信度 {high_conf} 条（≥80%）"
                + (f"，{need_review} 条需人工复查" if need_review else "")
            ),
        }))
    return evidence, events


async def _stream_stage5_risk(req: AnalyzeRequest, requirement, scored, loop) -> tuple:
    """Stage 5：风险评估 → (risks, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "risk", "status": "started"}))
    from .report_generator import _assess_risks
    risks = await loop.run_in_executor(None, _assess_risks, requirement, scored)
    events.append(_yield_sse("risk_done", {
        "status": "风险评估完成",
        "overall_risk_level": risks.overall_risk_level,
        "risk_count": len(risks.risk_items),
        "high": sum(1 for r in risks.risk_items if r.severity == "high"),
        "medium": sum(1 for r in risks.risk_items if r.severity == "medium"),
        "low": sum(1 for r in risks.risk_items if r.severity == "low"),
        "supply_summary": risks.supply_risk_summary,
        "engineering_summary": risks.engineering_risk_summary,
        "risk_items": [r.dict() for r in risks.risk_items],
    }))
    if req.thinking_depth != "off":
        high_c = sum(1 for r in risks.risk_items if r.severity == "high")
        med_c  = sum(1 for r in risks.risk_items if r.severity == "medium")
        low_c  = sum(1 for r in risks.risk_items if r.severity == "low")
        events.append(_yield_sse("thinking_delta", {
            "stage": "risk",
            "text": f"风险等级：{risks.overall_risk_level.upper()}（高 {high_c} / 中 {med_c} / 低 {low_c} 项）",
        }))
        if req.thinking_depth in ("contemplation", "exhaustive"):
            if risks.supply_risk_summary:
                events.append(_yield_sse("thinking_delta", {
                    "stage": "risk",
                    "text": f"供应链：{risks.supply_risk_summary}",
                }))
            if risks.engineering_risk_summary:
                events.append(_yield_sse("thinking_delta", {
                    "stage": "risk",
                    "text": f"工程设计：{risks.engineering_risk_summary}",
                }))
    return risks, events


async def _stream_stage6_report(req: AnalyzeRequest, requirement, scored, evidence, risks,
                                 t_start: float, loop, dual_model_enabled: bool = False) -> tuple:
    """Stage 6+7：报告生成 + 完成事件 → (events, report)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "report", "status": "started"}))
    from .report_generator import build_report
    report = await loop.run_in_executor(None, build_report, requirement, scored, evidence)
    sid = req.session_id or _DEFAULT_SESSION_ID
    _evict_session_stores()  # P0-4: prevent unbounded growth
    _session_reports[sid] = report
    _session_constraints[sid] = requirement

    # ── 参考设计获取（并行，Top-5 推荐器件）───────────────────────
    try:
        from .multi_agent import parallel_fetch_ref_designs_from_scored
        rd_map = await parallel_fetch_ref_designs_from_scored(report.recommended_parts or [])
        ref_designs: list = []
        for sp in (report.recommended_parts or [])[:5]:
            for rd in rd_map.get(sp.part.part_number, [])[:2]:
                ref_designs.append({
                    "part_number": sp.part.part_number,
                    "design_name": rd.get("name") or rd.get("title", "参考设计"),
                    "description": (rd.get("description") or "")[:200],
                    "url": rd.get("url") or rd.get("link", ""),
                })
        if ref_designs:
            report.reference_designs = ref_designs
            events.append(_yield_sse("reference_designs", {
                "designs": ref_designs,
                "count": len(ref_designs),
            }))
    except Exception:
        pass
    if report.summary_markdown:
        # Only send a brief natural-language summary, not the full markdown table
        recs = report.recommended_parts or []
        if recs:
            top = recs[0]
            pn = getattr(getattr(top, 'part', None), 'part_number', '?') or '?'
            score = int(getattr(getattr(top, 'score', None), 'total_score', 0) or 0)
            total = len(report.candidates or [])
            risk = getattr(risks, 'overall_risk_level', '').upper() or 'UNKNOWN'
            brief = (
                f"已完成选型分析，共找到 **{total}** 款候选器件。"
                f"首选推荐 **{pn}**（综合评分 {score} 分，整体风险 {risk}）。"
                f"完整评分明细和器件对比请在右侧面板查看。"
            )
            events.append(_yield_sse("text_delta", {"text": brief}))

    # ── 双模型验证（仅当用户已开启且管理员已配置验证模型时执行）──────
    try:
        from .dual_model_verify import verify_selection
        recs_for_verify = report.recommended_parts or []
        risk_level = getattr(risks, 'overall_risk_level', 'medium') or 'medium'
        top_scores = [float(getattr(getattr(sp, 'score', None), 'total_score', 0) or 0) for sp in recs_for_verify[:2]]
        score_delta_top2 = (top_scores[0] - top_scores[1]) if len(top_scores) >= 2 else 20.0
        should_verify = dual_model_enabled and bool(recs_for_verify) and not (risk_level == 'low' and score_delta_top2 > 10)
        if should_verify:
            primary_pn = getattr(getattr(recs_for_verify[0], 'part', None), 'part_number', '') or ''
            req_text = str(getattr(requirement, 'user_input', '') or '')
            cands_flat = []
            for sp in recs_for_verify[:8]:
                p = getattr(sp, 'part', None)
                s = getattr(sp, 'score', None)
                cands_flat.append({
                    "part_number": getattr(p, 'part_number', '?') or '?',
                    "manufacturer": getattr(p, 'manufacturer', '') or '',
                    "total_score": float(getattr(s, 'total_score', 0) or 0),
                })
            verify_result = await verify_selection(
                req_text, cands_flat, primary_pn, risk_level,
                primary_full_response=brief if 'brief' in dir() else "",
            )
            events.append(_yield_sse("dual_verify", {
                "passed": verify_result.passed,
                "primary_top": verify_result.primary_top,
                "verifier_top": verify_result.verifier_top,
                "agreement": verify_result.agreement,
                "score_delta": verify_result.score_delta,
                "needs_human_review": verify_result.needs_human_review,
                "notes": verify_result.verifier_notes,
                "verifier_full_response": verify_result.verifier_full_response,
                "judgment_reasoning": verify_result.judgment_reasoning,
            }))
            # Emit formatted comparison text when models disagree
            if not verify_result.agreement or verify_result.risk_conflict:
                comparison_text = (
                    "\n\n---\n**🔍 双模型验证结果（存在分歧）**\n\n"
                    f"**模型A（主模型）推荐：** {verify_result.primary_top}\n\n"
                    f"**模型B（验证模型）推荐：** {verify_result.verifier_top}\n\n"
                    f"**验证模型说明：** {verify_result.verifier_full_response}\n\n"
                )
                if verify_result.judgment_reasoning:
                    comparison_text += f"**裁判判断：** {verify_result.judgment_reasoning}\n"
                events.append(_yield_sse("text_delta", {"text": comparison_text}))
            else:
                events.append(_yield_sse("text_delta", {"text": f"\n\n✅ **双模型验证通过**：两个模型均推荐 **{verify_result.primary_top}**，结果一致。"}))
    except Exception:
        pass

    # ── 知识库自动写入（异步，不阻塞）──────────────────────────────
    try:
        from .kb_updater import auto_ingest_from_report
        await asyncio.to_thread(auto_ingest_from_report, report.dict())
    except Exception:
        pass
    elapsed = round(time.time() - t_start, 2)
    events.append(_yield_sse("done", {
        "status": "分析完成",
        "elapsed_s": elapsed,
        "request_id": report.request_id,
        "recommended_count": len(report.recommended_parts),
        "candidate_count": len(report.candidates),
        "overall_risk": risks.overall_risk_level,
        "summary": report.summary_markdown or "",
        "recommended_parts": [p.dict() for p in report.recommended_parts],
        "candidates": [p.dict() for p in report.candidates],
    }))
    return events, report


async def _stream_analyze(req: AnalyzeRequest, dual_model_enabled: bool = False) -> AsyncGenerator[str, None]:
    """异步生成器：按阶段推送 SSE 事件（含 B4 语义缓存集成）。

    各阶段已拆分为独立子函数：parse → search → score → evidence → risk → report。
    """
    t_start = time.time()
    loop = asyncio.get_running_loop()

    try:
        # ── B4：结构化选型缓存检查（必须先解析约束，禁止原始文本近邻命中）──
        from .semantic_cache import get_semantic_cache, canonical_constraint_fingerprint
        cache = get_semantic_cache()

        requirement, events = await _stream_stage1_parse(req, loop)
        for ev in events:
            yield ev

        fingerprint = canonical_constraint_fingerprint(requirement)
        cache_result = cache.get_exact(fingerprint) if fingerprint and not req.skip_cache else None
        cached_report = cache_result.get("cached_result", {}) if cache_result else {}
        restored_report = None
        if cache_result is not None:
            from .schemas import SelectionReport as _SelectionReport
            try:
                restored_report = _SelectionReport(**cached_report)
            except Exception:
                cache_result = None
                yield _yield_sse("cache_hit", {"hit": False, "reason": "invalid_cached_report"})

        if cache_result is not None and restored_report is not None:
            elapsed = round(time.time() - t_start, 2)
            yield _yield_sse("cache_hit", {"hit": True, "similarity": cache_result.get("similarity", 0)})
            sid = req.session_id or _DEFAULT_SESSION_ID
            _evict_session_stores()
            _session_reports[sid] = restored_report
            _session_constraints[sid] = restored_report.constraints

            scored = cached_report.get("candidates", []) or cached_report.get("recommended_parts", [])
            for i, s in enumerate(scored):
                part = s.get("part", {}) if isinstance(s, dict) else getattr(s, "part", None)
                score = s.get("score", {}) if isinstance(s, dict) else getattr(s, "score", None)
                if part:
                    pn = part.get("part_number", "") if isinstance(part, dict) else getattr(part, "part_number", "")
                    total_score = score.get("total_score", 0) if isinstance(score, dict) else (getattr(score, "total_score", 0) if score else 0)
                    rec_level = s.get("recommendation_level", "") if isinstance(s, dict) else getattr(s, "recommendation_level", "")
                    yield _yield_sse("score_update", {
                        "status": f"缓存命中: {pn}",
                        "index": i + 1, "total": len(scored),
                        "part_number": pn,
                        "total_score": total_score,
                        "recommendation_level": rec_level,
                    })

            summary = _rebuild_summary_from_cached_dict(cached_report)
            if summary:
                for line in summary.split('\n'):
                    if line.strip():
                        yield _yield_sse("text_delta", {"text": line})

            recs = cached_report.get("recommended_parts", [])
            all_candidates = cached_report.get("candidates", [])
            total_count = len(all_candidates)
            if recs:
                r0 = recs[0]
                part = r0.get("part", {}) if isinstance(r0, dict) else {}
                score = r0.get("score", {}) if isinstance(r0, dict) else {}
                pn = part.get("part_number", "?") if isinstance(part, dict) else "?"
                mfr = part.get("manufacturer", "") if isinstance(part, dict) else ""
                total_score = int(score.get("total_score", 0)) if isinstance(score, dict) else 0
                risks = cached_report.get("risks", {})
                risk_level = (risks.get("overall_risk_level", "?") if isinstance(risks, dict) else "?").upper()
                interaction_text = (
                    f"已为您找到 **{total_count}** 款候选器件，首选推荐 **{pn}**"
                    f"（{mfr}，综合评分 **{total_score} 分**）。"
                    "详细评分和对比请在右侧面板查看，也可以直接告诉我您的想法。"
                )
                yield _yield_sse("text_delta", {"text": interaction_text})

            rec_count = len(cached_report.get("recommended_parts", []))
            risk_val = cached_report.get("risks", {})
            if isinstance(risk_val, dict):
                overall = risk_val.get("overall_risk_level", "?")
            else:
                overall = getattr(risk_val, "overall_risk_level", "?")

            # 缓存命中路径此前跳过了 risk_done / evidence_done 事件，
            # 导致前端属性面板的"风险"/"证据"栏拿不到数据（即使候选列表已展示）。
            if isinstance(risk_val, dict) and risk_val:
                risk_items = risk_val.get("risk_items", []) or []
                yield _yield_sse("risk_done", {
                    "status": "风险评估完成（缓存）",
                    "overall_risk_level": risk_val.get("overall_risk_level", "low"),
                    "risk_count": len(risk_items),
                    "high": sum(1 for r in risk_items if (r or {}).get("severity") == "high"),
                    "medium": sum(1 for r in risk_items if (r or {}).get("severity") == "medium"),
                    "low": sum(1 for r in risk_items if (r or {}).get("severity") == "low"),
                    "supply_summary": risk_val.get("supply_risk_summary"),
                    "engineering_summary": risk_val.get("engineering_risk_summary"),
                    "risk_items": risk_items,
                })

            cached_evidence = cached_report.get("evidence", []) or []
            evidence_items = []
            for item in cached_evidence:
                if not isinstance(item, dict):
                    continue
                evidence_items.append({
                    "part_number": item.get("part_number", ""),
                    "claim": item.get("claim", ""),
                    "evidence_type": item.get("evidence_type", ""),
                    "source_field": item.get("source_field", ""),
                    "confidence": item.get("confidence", 0.0),
                    "need_human_review": item.get("need_human_review", False),
                })
            avg_confidence = (
                round(sum(float(item.get("confidence", 0) or 0) for item in evidence_items) / len(evidence_items), 3)
                if evidence_items else 0.0
            )
            yield _yield_sse("evidence_done", {
                "status": "证据链构建完成（缓存）",
                "evidence_count": len(evidence_items),
                "avg_confidence": avg_confidence,
                "evidence_items": evidence_items,
            })

            yield _yield_sse("done", {
                "status": "分析完成（语义缓存命中）",
                "elapsed_s": elapsed,
                "request_id": cached_report.get("request_id", "cached"),
                "recommended_count": rec_count,
                "candidate_count": len(cached_report.get("candidates") or cached_report.get("scored_parts") or []),
                "overall_risk": overall,
                "summary": summary,
                "cache_hit": True,
                "recommended_parts": cached_report.get("recommended_parts", []),
                "candidates": (cached_report.get("candidates") or
                               cached_report.get("scored_parts") or []),
            })
            return

        yield _yield_sse("cache_hit", {"hit": False})

        # ── Stage 2：器件搜索 ──────────────────────────────────
        candidates, events = await _stream_stage2_search(req, requirement, loop)
        for ev in events:
            yield ev

        # ── Stage 3：评分计算 ──────────────────────────────────
        scored, events = await _stream_stage3_score(req, requirement, candidates, loop)
        for ev in events:
            yield ev

        # ── 生命周期告警（并发查询替代料）────────────────────────
        lc_alerts = []
        for sp in scored[:10]:
            lc = (sp.part.lifecycle_status or "").lower().strip()
            if lc in ("nrnd", "ltb", "eol", "obsolete", "discontinued"):
                severity = "HIGH" if lc in ("eol", "obsolete", "discontinued") else "MEDIUM"
                lc_alerts.append({
                    "part_number": sp.part.part_number,
                    "manufacturer": sp.part.manufacturer or "",
                    "lifecycle_status": sp.part.lifecycle_status,
                    "severity": severity,
                    "alternatives": [],
                })
        if lc_alerts:
            from .multi_agent import parallel_lifecycle_replacements
            lc_alerts = await parallel_lifecycle_replacements(lc_alerts)
            yield _yield_sse("lifecycle_alert", {
                "alerts": lc_alerts,
                "count": len(lc_alerts),
                "has_high": any(a["severity"] == "HIGH" for a in lc_alerts),
            })
            if req.thinking_depth != "off":
                high_parts = [a["part_number"] for a in lc_alerts if a["severity"] == "HIGH"]
                if high_parts:
                    yield _yield_sse("thinking_delta", {
                        "stage": "lifecycle",
                        "text": f"⚠ 发现 {len(high_parts)} 个已停产/EOL 器件：{', '.join(high_parts)}，已并发查询替代料",
                    })

        # ── Stages 4+5：证据构建 + 风险评估（并发）────────────────
        yield _yield_sse("agent_activity", {
            "agent": "AnalysisOrchestrator",
            "status": "EvidenceAgent + RiskAgent 并发运行",
            "phase": "analysis",
        })
        (evidence, ev4), (risks, ev5) = await asyncio.gather(
            _stream_stage4_evidence(req, scored, requirement, loop),
            _stream_stage5_risk(req, requirement, scored, loop),
        )
        for ev in ev4:
            yield ev
        for ev in ev5:
            yield ev

        # ── P1/P5: CriticNode 自省检查 ─────────────────────
        from .langgraph_orchestrator import critic_node
        critic_state = {"constraints": requirement, "scored": scored, "retry_count": 0}
        critic_result = critic_node(critic_state)
        if not critic_result.get("critic_passed"):
            yield _yield_sse("warning", {"message": f"Critic: {critic_result.get('error', '')}"})

        # ── Stage 6：报告生成 + 完成 ───────────────────────────
        _model_ver = get_active_model()
        events, report = await _stream_stage6_report(req, requirement, scored, evidence, risks, t_start, loop, dual_model_enabled=dual_model_enabled)
        # P2-7: bind model version and timestamp for reproducibility
        try:
            import datetime as _dt
            report.model_version = _model_ver
            report.generated_at = _dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        except Exception:
            pass
        for ev in events:
            yield ev

        # ── B4：写入版本化、结构化选型缓存 ─────────────────────────────
        try:
            from .semantic_cache import get_semantic_cache, canonical_constraint_fingerprint
            _report_dict = report.dict()
            _fingerprint = canonical_constraint_fingerprint(report.constraints)
            if _fingerprint:
                get_semantic_cache().set_exact(_fingerprint, _report_dict)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"语义缓存写入失败: {e}")

        # ── Memory：记录选型历史 ────────────────────────────────
        try:
            from .memory import record_selection
            recs = report.recommended_parts
            summary_text = f"推荐 {len(recs)} 款" + (f", Top1={recs[0].part.part_number}" if recs else ", 无推荐")
            record_selection(req.user_input, summary_text)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"选型历史记录失败: {e}")

    except Exception as exc:
        elapsed = round(time.time() - t_start, 2)
        yield _yield_sse("error", {
            "status": "错误",
            "message": str(exc),
            "elapsed_s": elapsed,
            "traceback": traceback.format_exc()[:500],
        })


@app.post("/analyze/stream")
async def analyze_stream_endpoint(
    body: AnalyzeRequest,
    current_user=Depends(get_current_user),
):
    """流式输出端点：SSE 逐段推送选型报告

    ── B4：语义缓存支持 ──
    - 响应头 X-Cache: HIT 表示命中缓存
    - 响应头 X-Cache: MISS 表示未命中缓存

    示例事件流：
    - cache_hit: 缓存命中状态（HIT/MISS）
    - parse_done: 需求解析完成
    - search_done: 搜索完成 + 候选数量
    - score_update: 评分完成（多次）
    - evidence_done: 证据链完成
    - risk_done: 风险评估完成 + RiskIR
    - text_delta: 报告文本片段（多次）
    - done: 完成 + 总耗时 + 完整报告
    """
    # The generator emits the authoritative cache_hit event after parsing the
    # structured requirement. A raw-text header would be misleading.
    cache_header = "DEFERRED"

    return StreamingResponse(
        _stream_analyze(body, dual_model_enabled=bool(getattr(current_user, 'dual_model_enabled', False))),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
            "X-Cache": cache_header,
        }
    )


# ── 电路图生成端点（B2 任务）────────────────────────────────────────

@app.get("/schematic/{topology}")
async def get_schematic(topology: str, Vin: float, Vout: float, Iout: float):
    """生成参数化应用电路 SVG

    Args:
        topology: 拓扑类型 ('buck', 'boost', 'ldo')
        Vin: 输入电压 (V)
        Vout: 输出电压 (V)
        Iout: 输出电流 (A)

    Returns:
        SVG 格式的电路图
    """
    try:
        from .schematic_generator import generate_schematic
        svg = generate_schematic(topology, Vin, Vout, Iout)
        return Response(content=svg, media_type="image/svg+xml")
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"detail": str(e)},
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"detail": f"电路图生成错误: {str(e)}"},
        )


# ── 器件选择端点 ────────────────────────────────────────────────

class SelectPartRequest(BaseModel):
    session_id: str = _DEFAULT_SESSION_ID
    part_number: str

@app.post("/select-part")
async def select_part_endpoint(body: SelectPartRequest, current_user=Depends(get_current_user)):
    """用户选择具体器件后，记录选定器件并生成其专属报告。"""
    report = _session_reports.get(body.session_id)
    if not report:
        return JSONResponse(status_code=404, content={"detail": "未找到该会话的选型报告"})

    # 验证器件是否在候选列表中
    candidate_parts = _report_parts(report)
    all_part_numbers = [_part_number(sp) for sp in candidate_parts]
    if body.part_number not in all_part_numbers:
        return JSONResponse(status_code=400, content={"detail": f"器件 {body.part_number} 不在本次选型结果中"})

    _session_selected_part[body.session_id] = body.part_number
    return {"status": "ok", "part_number": body.part_number}


class InterpretSelectionRequest(BaseModel):
    session_id: str = _DEFAULT_SESSION_ID
    user_input: str

@app.post("/interpret-selection")
async def interpret_selection_endpoint(body: InterpretSelectionRequest, current_user=Depends(get_current_user)):
    """用 LLM 理解用户输入是否在选取器件，若是则返回对应型号。"""
    report = _session_reports.get(body.session_id)
    if not report:
        return {"selected": None}

    # 收集候选器件列表
    candidate_parts = _report_parts(report)
    all_parts = []
    for sp in candidate_parts:
        pn = _part_number(sp)
        if isinstance(sp, dict):
            part = sp.get("part") or {}
            mfr = part.get("manufacturer", "") if isinstance(part, dict) else ""
            score_data = sp.get("score") or {}
            score = int(score_data.get("total_score", 0)) if isinstance(score_data, dict) else 0
        else:
            mfr = getattr(getattr(sp, "part", None), "manufacturer", "") or ""
            score = int(getattr(getattr(sp, "score", None), "total_score", 0) or 0)
        level = sp.get("recommendation_level", "") if isinstance(sp, dict) else getattr(sp, "recommendation_level", "")
        all_parts.append(f"{pn} ({mfr}, 评分{score}, {level})")

    if not all_parts:
        return {"selected": None}

    part_list_str = "\n".join(f"{i+1}. {p}" for i, p in enumerate(all_parts))
    prompt = (
        f"用户输入：{body.user_input}\n\n"
        f"候选器件列表：\n{part_list_str}\n\n"
        "请判断用户是否在从上述列表中选择某个器件。如果是，只返回该器件的型号（MPN），不要其他内容。如果用户不是在选择器件（例如在提问、调整需求、闲聊等），只返回 null。"
    )

    try:
        from .llm_client import call_openai_chat_text
        result = call_openai_chat_text(
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
        )
        result = result.strip().strip('"\'`')
        # 检查是否在列表中
        all_pns = {_part_number(sp) for sp in candidate_parts}
        if result in all_pns:
            return {"selected": result}
    except Exception:
        pass
    return {"selected": None}


# ── 三类报告 Markdown 端点 ───────────────────────────────────────

@app.get("/report/{report_type}")
async def get_report(report_type: str, session_id: Optional[str] = None, current_user=Depends(get_current_user)):
    """返回指定会话的三类报告 Markdown 内容。

    Args:
        report_type: 'bom' | 'risk' | 'topology'
        session_id: 可选会话 ID（未提供时使用默认会话）

    Returns:
        {"content": "Markdown文本", "type": "bom|risk|topology"}
    """
    sid = session_id or _DEFAULT_SESSION_ID
    selected_pn = _session_selected_part.get(sid)
    if not selected_pn:
        return JSONResponse(status_code=400, content={"detail": "请先选择具体器件（回复编号如 1、2 等）后再查看报告"})

    _latest_report = _session_reports.get(sid)
    _latest_constraints = _session_constraints.get(sid)
    if _latest_report is None:
        return JSONResponse(status_code=404, content={"detail": "暂无分析报告，请先执行一次选型分析"})

    try:
        from .output_generator import generate_all_reports
        from .output_bom import generate_bom
        from .output_generator import generate_risk_report, generate_topology
        import copy

        # 构建仅含选定器件的报告副本
        filtered_report = copy.deepcopy(_latest_report)
        filtered_report.recommended_parts = [
            sp for sp in filtered_report.recommended_parts
            if sp.part.part_number == selected_pn
        ]
        filtered_report.candidates = [
            sp for sp in filtered_report.candidates
            if sp.part.part_number == selected_pn
        ]
        if filtered_report.evidence:
            filtered_report.evidence = [
                e for e in filtered_report.evidence
                if e.part_number == selected_pn
            ]

        rag_context = ""
        try:
            from .output_generator import _extract_rag_context
            rag_context = _extract_rag_context(filtered_report)
        except Exception:
            pass

        if report_type == "bom":
            md = generate_bom(filtered_report, rag_context=rag_context)
        elif report_type == "risk":
            md = generate_risk_report(filtered_report, _latest_constraints, rag_context=rag_context)
        elif report_type == "topology":
            md = generate_topology(_latest_constraints, filtered_report, rag_context=rag_context)
        else:
            return JSONResponse(status_code=400, content={"detail": f"未知报告类型: {report_type}，支持 bom/risk/topology"})

        return {"content": md, "type": report_type}
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"报告生成错误: {str(e)}"})


# ── 文件上传/解析端点 ──────────────────────────────────────────

ALLOWED_EXTENSIONS = {".pdf", ".csv", ".txt", ".md", ".json", ".xlsx", ".xls"}

@app.post("/upload/parse")
async def upload_and_parse(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    """上传文件并提取文本内容（供 LLM 解析）。

    支持: PDF (数据手册), CSV (BOM表), TXT/MD, JSON, Excel
    """
    if not file.filename:
        return JSONResponse(status_code=400, content={"detail": "未提供文件"})

    ext = file.filename.lower().rsplit(".", 1)[-1] if "." in file.filename else ""
    if f".{ext}" not in ALLOWED_EXTENSIONS:
        return JSONResponse(status_code=400,
            content={"detail": f"不支持的文件类型: .{ext}，支持: {', '.join(ALLOWED_EXTENSIONS)}"})

    try:
        content_bytes = await file.read()

        if ext in ("txt", "md", "json", "csv"):
            content = content_bytes.decode("utf-8", errors="replace")
        elif ext == "pdf":
            try:
                import io
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(content_bytes))
                pages = [p.extract_text() or "" for p in reader.pages[:30]]  # 前30页（数据手册通常40-100页）
                content = "\n\n".join(pages)
            except ImportError:
                content = f"[PDF文件: {file.filename}] — PyPDF2 未安装，无法提取文本"
        elif ext in ("xlsx", "xls"):
            try:
                import io
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True)
                sheets = []
                total_rows = 0
                max_total_rows = 10000  # 总行数上限，防止内存溢出
                for name in wb.sheetnames:  # 读取所有 Sheet
                    ws = wb[name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        if total_rows >= max_total_rows:
                            break
                        rows.append("\t".join(str(c) if c else "" for c in row))
                        total_rows += 1
                    sheets.append(f"=== {name} ===\n" + "\n".join(rows))
                    if total_rows >= max_total_rows:
                        break
                content = "\n\n".join(sheets)
            except ImportError:
                content = f"[Excel文件: {file.filename}] — openpyxl 未安装"
        else:
            content = f"[不支持的文件类型: {file.filename}]"

        return {
            "filename": file.filename,
            "type": ext,
            "size_bytes": len(content_bytes),
            "content": content[:5000],  # 限制返回大小
            "truncated": len(content) > 5000,
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"文件解析错误: {str(e)}"})


# ── BOM 健康度诊断端点（B5）──────────────────────────────────────

@app.post("/bom/validate")
async def bom_validate_endpoint(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    """上传现有 BOM (CSV/Excel) → eZ-PLM 逐条验证 → 生命周期 + 替代料报告。"""
    if not file.filename:
        return JSONResponse(status_code=400, content={"detail": "未提供文件"})
    content_bytes = await file.read()
    ext = file.filename.lower().rsplit(".", 1)[-1] if "." in file.filename else ""

    mpns: list = []
    try:
        if ext == "csv":
            import csv as _csv, io as _io
            text = content_bytes.decode("utf-8", errors="replace")
            reader = _csv.DictReader(_io.StringIO(text))
            mpn_col = None
            for row in reader:
                if mpn_col is None:
                    _MPN_HEADERS = {"mpn", "part number", "partnumber", "型号", "物料号", "料号", "part no.", "part no", "mfr pn", "mfr part"}
                    mpn_col = next((k for k in row if k.strip().lower() in _MPN_HEADERS), list(row.keys())[0] if row else None)
                val = (row.get(mpn_col) or "").strip()
                if val:
                    mpns.append(val)
        elif ext in ("xlsx", "xls"):
            import io as _io, openpyxl as _xl
            wb = _xl.load_workbook(_io.BytesIO(content_bytes), read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                return JSONResponse(status_code=400, content={"detail": "Excel 文件为空"})
            hdr = [str(h).strip().lower() if h else "" for h in header_row]
            _MPN_HEADERS = {"mpn", "part number", "partnumber", "型号", "物料号", "料号", "part no.", "part no"}
            col_idx = next((i for i, h in enumerate(hdr) if h in _MPN_HEADERS), 0)
            for row in rows_iter:
                if row and col_idx < len(row) and row[col_idx] and str(row[col_idx]).strip():
                    mpns.append(str(row[col_idx]).strip())
        else:
            return JSONResponse(status_code=400, content={"detail": "仅支持 CSV / Excel (.xlsx/.xls) 格式"})
    except Exception as e:
        return JSONResponse(status_code=400, content={"detail": f"文件解析失败: {str(e)}"})

    mpns = list(dict.fromkeys(mpns))[:100]  # 去重 + 限制 100 条
    if not mpns:
        return JSONResponse(status_code=400, content={"detail": "未能提取 MPN 列表，请确认列头包含 MPN / 型号 / Part Number 等"})

    from .ezplm_client import search_part_by_mpn as _mpn_q, find_replacements as _find_alt
    from .output_bom import _lifecycle_normalized, _lifecycle_cn
    import asyncio as _aio

    loop = _aio.get_running_loop()
    results, issues = [], []
    summary = {"active": 0, "nrnd": 0, "eol_obsolete": 0, "unknown": 0, "not_found": 0, "domestic_count": 0}

    for mpn in mpns:
        part = await loop.run_in_executor(None, _mpn_q, mpn)
        if not part:
            summary["not_found"] += 1
            issues.append({"mpn": mpn, "found": False, "lifecycle_status": "Unknown",
                           "risk_level": "LOW", "message": f"{mpn} 在 eZ-PLM 中未找到", "alternatives": []})
            continue

        lc_raw = _lifecycle_normalized(part.lifecycle_status)
        if lc_raw == "Active":
            summary["active"] += 1
        elif lc_raw == "NRND":
            summary["nrnd"] += 1
        elif lc_raw in ("EOL", "Obsolete", "LTB"):
            summary["eol_obsolete"] += 1
        else:
            summary["unknown"] += 1
        if getattr(part, "is_domestic", False):
            summary["domestic_count"] += 1

        alternatives, risk_level = [], "LOW"
        if lc_raw in ("EOL", "Obsolete"):
            risk_level = "HIGH"
            try:
                alts = await loop.run_in_executor(None, _find_alt, mpn)
                alternatives = [a.part_number for a in (alts or [])[:3]]
            except Exception:
                pass
        elif lc_raw in ("NRND", "LTB"):
            risk_level = "MEDIUM"

        item = {
            "mpn": mpn, "found": True,
            "manufacturer": part.manufacturer or "",
            "lifecycle_status": lc_raw,
            "lifecycle_cn": _lifecycle_cn(part.lifecycle_status),
            "risk_level": risk_level,
            "is_domestic": getattr(part, "is_domestic", False),
            "package": part.package or "",
            "alternatives": alternatives,
        }
        results.append(item)
        if risk_level != "LOW":
            issues.append(item)

    total = len(mpns)
    return {
        "total": total,
        "checked": len(results),
        "issues_count": len(issues),
        "issues": issues,
        "all_results": results,
        "summary": {**summary, "domestic_ratio": round(summary["domestic_count"] / total, 3) if total else 0},
        "health_score": round(summary["active"] / total * 100, 1) if total else 0,
    }


# ── BOM Excel 导出端点（B6）────────────────────────────────────

@app.post("/export/bom")
async def export_bom_endpoint(session_id: Optional[str] = None, current_user=Depends(get_current_user)):
    """导出工程级 BOM Excel 文件（三 Sheet）。

    Args:
        session_id: 可选会话 ID（未提供时使用默认会话）

    Returns:
        .xlsx binary (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)
    """
    sid = session_id or _DEFAULT_SESSION_ID
    _latest_report = _session_reports.get(sid)
    if _latest_report is None:
        return JSONResponse(status_code=404, content={"detail": "暂无分析报告，请先执行一次选型分析"})

    try:
        from .output_bom import generate_bom_excel
        xlsx_bytes = generate_bom_excel(_latest_report)
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=BOM_{_latest_report.request_id[:8]}.xlsx",
            },
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"Excel 生成错误: {str(e)}"})


@app.post("/export/decision-package")
async def export_decision_package_endpoint(session_id: Optional[str] = None, current_user=Depends(get_current_user)):
    """导出选型决策包（IEC/IATF 格式，含国产化分析，7 Sheet Excel）。"""
    sid = session_id or _DEFAULT_SESSION_ID
    report = _session_reports.get(sid)
    if report is None:
        return JSONResponse(status_code=404, content={"detail": "暂无分析报告，请先执行一次选型分析"})
    try:
        from .output_decision_package import generate_decision_package
        xlsx = generate_decision_package(report)
        rid = getattr(report, 'request_id', 'pkg')[:8]
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=DecisionPackage_{rid}.xlsx"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"生成错误: {str(e)}"})


# ── 评分维度重算（支持用户开关评分因素）────────────────────
DIMENSION_MAP = {
    "parameter_match": "parameter_match_score",
    "supply_risk": "supply_risk_score",
    "cost": "cost_score",
    "domestic": "domestic_score",
    "evidence": "evidence_score",
}

DIMENSION_LABELS = {
    "parameter_match": "参数匹配度",
    "supply_risk": "供应链风险",
    "cost": "成本",
    "domestic": "国产化",
    "evidence": "证据可信度",
}

class RecalculateRequest(BaseModel):
    session_id: str = _DEFAULT_SESSION_ID
    dimensions: list[str] = list(DIMENSION_MAP.keys())  # 用户启用的维度列表

@app.post("/recalculate")
async def recalculate_endpoint(body: RecalculateRequest, current_user=Depends(get_current_user)):
    """根据用户启用的评分维度重新计算推荐分数。"""
    report = _session_reports.get(body.session_id)
    if not report:
        return JSONResponse(status_code=404, content={"detail": "未找到该会话的报告"})

    if not body.dimensions:
        return JSONResponse(status_code=400, content={"detail": "至少需要一个评分维度"})

    # 校验所有维度名称
    for d in body.dimensions:
        if d not in DIMENSION_MAP:
            return JSONResponse(status_code=400, content={"detail": f"未知维度: {d}"})

    updated_parts = []
    for sp in _report_parts(report):
        score = sp.score
        vals = []
        for d in body.dimensions:
            v = getattr(score, DIMENSION_MAP[d], None)
            if v is not None:
                vals.append(v)
        # 重算 total_score：启用维度的均值
        new_total = round(sum(vals) / len(vals), 2) if vals else 0.0
        updated_parts.append({
            "part_number": sp.part.part_number,
            "manufacturer": sp.part.manufacturer,
            "scores": {d: getattr(score, DIMENSION_MAP[d], 0) for d in DIMENSION_MAP},
            "active_dimensions": body.dimensions,
            "total_score": new_total,
        })

    return {
        "parts": sorted(updated_parts, key=lambda x: x["total_score"], reverse=True),
        "active_dimensions": body.dimensions,
    }


# ── 工作流 AI 生成 ─────────────────────────────────────────────────────
class WorkflowGenerateRequest(BaseModel):
    description: str

_WORKFLOW_ICONS = ["Brain", "Search", "MessageSquare", "FileText", "BarChart2",
                   "CheckCircle", "Filter", "Cpu", "Zap", "Shield", "GitMerge",
                   "Radio", "AlertTriangle"]
_WORKFLOW_COLORS = ["bg-violet-100 text-violet-700", "bg-teal-100 text-teal-700",
                    "bg-blue-100 text-blue-700", "bg-amber-100 text-amber-700",
                    "bg-rose-100 text-rose-700", "bg-emerald-100 text-emerald-700"]

_WORKFLOW_SYSTEM = """你是一个工作流设计专家。根据用户对业务流程的自然语言描述，生成一个有向工作流图。

要求：
1. 返回纯 JSON，不含 markdown 代码块。
2. JSON 结构：{"name": "...", "nodes": [...], "edges": [...]}
3. 每个 node 字段：id(字符串), iconName(从列表选), label(节点名≤8字), color(从列表选), description(≤20字), x(数字), y(数字)
4. iconName 只能从以下选择：Brain Search MessageSquare FileText BarChart2 CheckCircle Filter Cpu Zap Shield GitMerge Radio AlertTriangle
5. color 只能从以下选择：bg-violet-100 text-violet-700 / bg-teal-100 text-teal-700 / bg-blue-100 text-blue-700 / bg-amber-100 text-amber-700 / bg-rose-100 text-rose-700 / bg-emerald-100 text-emerald-700
6. 节点 x/y 坐标：从左到右或从上到下布局，x 步长约 220，y 步长约 120，起点 (80, 80)
7. 每个 edge 字段：source(节点id), target(节点id)
8. 节点数量 4-10 个，确保逻辑连贯、覆盖描述中的关键步骤"""

@app.post("/workflow/generate")
async def workflow_generate(body: WorkflowGenerateRequest, current_user=Depends(get_current_user)):
    from .llm_client import call_openai_chat
    import json as _json

    messages = [
        {"role": "system", "content": _WORKFLOW_SYSTEM},
        {"role": "user", "content": f"请根据以下业务流程描述，设计工作流图：\n\n{body.description}"},
    ]
    try:
        result = call_openai_chat(messages, temperature=0.3, thinking_depth="off")
        raw = result.get("content", "")
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("```", 2)[1]
            if cleaned.startswith("json"):
                cleaned = cleaned[4:]
        data = _json.loads(cleaned.strip())
        return data
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"AI 生成失败: {str(e)}"})


# ── 前端静态文件托管（单端口部署模式）─────────────────────
import os as _os
_static_dir = _os.path.join(_os.path.dirname(__file__), "static")
if _os.path.exists(_static_dir):
    from fastapi.staticfiles import StaticFiles
    from fastapi.responses import HTMLResponse
    app.mount("/", StaticFiles(directory=_static_dir, html=True), name="static")

    @app.get("/setup", response_class=HTMLResponse)
    async def setup_page():
        """Onboarding 配置向导（代理到首页，前端 SPA 处理路由）。"""
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/")
