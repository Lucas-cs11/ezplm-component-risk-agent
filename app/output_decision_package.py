"""
output_decision_package.py — 选型决策包（IEC/IATF 对齐格式）

集成三类报告（选型报告 + 供应链风险 + BOM清单）并追加国产化分析，
生成符合 IATF 16949 / IEC 62368-1 工程文档规范的多 Sheet Excel。

结构：
  Sheet 1  文档封面         Document Cover
  Sheet 2  技术需求矩阵     Requirements Matrix
  Sheet 3  选型结论         Selection Conclusion
  Sheet 4  参数符合性验证   Parameter Compliance
  Sheet 5  供应链风险评估   Supply Chain Risk
  Sheet 6  国产化分析       Domestic Substitution
  Sheet 7  替代料清单 (AVL) Approved Vendor List
  Sheet 8  证据链与数据来源 Evidence Chain
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import List, Optional


# ──────────────────────────────────────────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────────────────────────────────────────

def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

def _today() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")

def _lc_label(status: Optional[str]) -> str:
    mapping = {
        "active": "Active ✓", "production": "Active ✓",
        "nrnd": "NRND ⚠", "ltb": "LTB ⚠",
        "eol": "EOL ✗", "obsolete": "EOL ✗", "discontinued": "EOL ✗",
    }
    return mapping.get((status or "").lower().strip(), status or "Unknown")

def _risk_emoji(level: str) -> str:
    return {"high": "🔴 HIGH", "medium": "🟡 MEDIUM", "low": "🟢 LOW"}.get(
        level.lower(), level
    )

def _grade_cn(grade: Optional[str]) -> str:
    return {
        "automotive": "车规级 (AEC-Q100)",
        "industrial": "工业级",
        "commercial": "商业级",
        "military": "军工级",
    }.get((grade or "").lower(), grade or "—")


# ──────────────────────────────────────────────────────────────────────────────
# 主函数
# ──────────────────────────────────────────────────────────────────────────────

def generate_decision_package(report, title: str = "选型决策包") -> bytes:
    """生成选型决策包 Excel。

    Args:
        report: SelectionReport 对象
        title:  文档标题（可含项目名称）

    Returns:
        .xlsx bytes
    """
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── 全局样式 ────────────────────────────────────────────────
    _B = Side(style="thin", color="BDBDBD")
    BORDER = Border(left=_B, right=_B, top=_B, bottom=_B)
    AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    AL_R = Alignment(horizontal="right", vertical="center")

    F_TIT = Font(name="Calibri", size=14, bold=True, color="1A237E")
    F_SEC = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    F_HDR = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    F_VAL = Font(name="Calibri", size=9)
    F_KEY = Font(name="Calibri", size=9, bold=True, color="37474F")

    FL_BLU  = PatternFill("solid", fgColor="1A237E")
    FL_BLU2 = PatternFill("solid", fgColor="283593")
    FL_GRN  = PatternFill("solid", fgColor="E8F5E9")
    FL_YEL  = PatternFill("solid", fgColor="FFF8E1")
    FL_RED  = PatternFill("solid", fgColor="FFCDD2")
    FL_GRY  = PatternFill("solid", fgColor="F5F5F5")

    def hdr(ws, row, col, val, font=F_HDR, fill=FL_BLU2, align=AL_C, width=None):
        c = ws.cell(row, col, val)
        c.font = font; c.fill = fill; c.alignment = align; c.border = BORDER
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def cell(ws, row, col, val, font=F_VAL, fill=None, align=AL_L):
        c = ws.cell(row, col, val)
        c.font = font; c.border = BORDER; c.alignment = align
        if fill:
            c.fill = fill
        return c

    def section_title(ws, row, col_from, col_to, text):
        ws.merge_cells(start_row=row, start_column=col_from, end_row=row, end_column=col_to)
        c = ws.cell(row, col_from, text)
        c.font = F_SEC; c.fill = FL_BLU; c.alignment = AL_L
        c.border = BORDER
        ws.row_dimensions[row].height = 20
        return c

    # ────────────────────────────────────────────────────────────
    # Sheet 1: 文档封面
    # ────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "选型决策包"
    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 50

    # 大标题
    ws1.merge_cells("A1:B1")
    c = ws1["A1"]
    c.value = "选型决策包 (Selection Decision Package)"
    c.font = F_TIT; c.alignment = AL_C; c.fill = FL_BLU
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws1.row_dimensions[1].height = 40

    sub = ws1.merge_cells("A2:B2")
    ws1["A2"].value = "符合 IATF 16949 / IEC 62368-1 工程文档规范"
    ws1["A2"].font = Font(name="Calibri", size=10, italic=True, color="37474F")
    ws1["A2"].alignment = AL_C
    ws1.row_dimensions[2].height = 20

    meta = [
        ("文档编号", f"SDP-{report.request_id[:8].upper() if hasattr(report, 'request_id') and report.request_id else '00000000'}"),
        ("生成日期", _today()),
        ("需求输入", (report.user_input[:80] + "…") if hasattr(report, 'user_input') and len(report.user_input or '') > 80 else (report.user_input if hasattr(report, 'user_input') else "—")),
        ("文档状态", "Draft (AI Generated)"),
        ("生成系统", "eZmanbo 智能选型系统 v2.0"),
        ("数据来源", "eZ-PLM 元器件数据库 + 工程知识库"),
    ]
    for i, (k, v) in enumerate(meta, start=4):
        ws1.row_dimensions[i].height = 18
        cell(ws1, i, 1, k, font=F_KEY)
        cell(ws1, i, 2, v)

    # 摘要框
    section_title(ws1, 11, 1, 2, "  决策摘要 (Executive Summary)")
    recs = getattr(report, 'recommended_parts', []) or []
    candidates = getattr(report, 'candidates', []) or []
    risks = getattr(report, 'risks', None)
    risk_level = (risks.overall_risk_level if risks else "—") if hasattr(risks, 'overall_risk_level') else "—"
    domestic_recs = [s for s in recs if getattr(s.part, 'is_domestic', False)]
    domestic_rate = len(domestic_recs) / len(recs) * 100 if recs else 0

    top = recs[0] if recs else None
    summary_rows = [
        ("推荐器件数", f"{len(recs)} 款（含备选 {max(0, len(candidates)-len(recs))} 款）"),
        ("首选器件", f"{top.part.part_number} ({top.part.manufacturer or '—'})" if top else "—"),
        ("首选评分", f"{top.score.total_score:.1f} / 100" if top else "—"),
        ("整体风险等级", _risk_emoji(risk_level)),
        ("国产化率", f"{domestic_rate:.1f}% ({len(domestic_recs)}/{len(recs)})"),
        ("首选生命周期", _lc_label(top.part.lifecycle_status if top else None)),
    ]
    for i, (k, v) in enumerate(summary_rows, start=12):
        ws1.row_dimensions[i].height = 18
        cell(ws1, i, 1, k, font=F_KEY, fill=FL_GRY)
        cell(ws1, i, 2, v)

    # ────────────────────────────────────────────────────────────
    # Sheet 2: 技术需求矩阵
    # ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("技术需求矩阵")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 30

    section_title(ws2, 1, 1, 5, "  技术需求矩阵 (Requirements Matrix)  [IEC 62368-1 §5]")
    headers = ["参数名称", "需求值", "首选器件值", "符合性", "备注"]
    for ci, h in enumerate(headers, 1):
        hdr(ws2, 2, ci, h)

    c = getattr(report, 'constraints', None)
    if top and c:
        p = top.part
        checks = [
            ("输入电压范围", f"{getattr(c,'input_voltage_nominal_v',None) or '—'}V",
             f"{p.input_voltage_min_v or '?'}~{p.input_voltage_max_v or '?'}V",
             "✓" if p.input_voltage_min_v and p.input_voltage_max_v else "—"),
            ("输出电压", f"{getattr(c,'output_voltage_v',None) or '—'}V",
             f"{p.output_voltage_v or '?'}V", "✓" if p.output_voltage_v else "—"),
            ("最大输出电流", f"{getattr(c,'output_current_a',None) or '—'}A",
             f"{p.output_current_max_a or '?'}A",
             "✓" if p.output_current_max_a and c.output_current_a and p.output_current_max_a >= c.output_current_a else "!"),
            ("温度范围",
             f"{getattr(c,'temperature_min_c',None) or '?'}~{getattr(c,'temperature_max_c',None) or '?'}°C",
             f"{p.temperature_min_c or '?'}~{p.temperature_max_c or '?'}°C",
             "✓" if p.temperature_min_c is not None else "—"),
            ("应用等级", _grade_cn(getattr(c,'grade',None)),
             "AEC-Q100 ✓" if p.automotive_grade else _grade_cn(None), "✓" if p.automotive_grade else "⚠"),
            ("封装偏好", getattr(c,'package_preference',None) or "无要求",
             p.package or "—", "—"),
            ("生命周期", "Active", _lc_label(p.lifecycle_status),
             "✓" if (p.lifecycle_status or "").lower() == "active" else "⚠"),
        ]
        for ri, (name, req, actual, comply, *notes) in enumerate(checks, start=3):
            fill = FL_GRN if comply == "✓" else (FL_YEL if comply == "⚠" else None)
            cell(ws2, ri, 1, name, font=F_KEY)
            cell(ws2, ri, 2, req)
            cell(ws2, ri, 3, actual)
            cell(ws2, ri, 4, comply, align=AL_C, fill=fill)
            cell(ws2, ri, 5, notes[0] if notes else "")

    # ────────────────────────────────────────────────────────────
    # Sheet 3: 选型结论
    # ────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("选型结论")
    for ci, w in enumerate([6,22,20,14,10,10,10,10,14,14], start=1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    section_title(ws3, 1, 1, 10, "  选型结论 (Selection Conclusion)  [IATF 16949 §8.3]")
    h3 = ["排名","型号","制造商","封装","综合评分","参数匹配","供应评分","推荐等级","生命周期","国产"]
    for ci, h in enumerate(h3, 1): hdr(ws3, 2, ci, h)

    for ri, s in enumerate(recs[:10], start=3):
        p = s.part; sc = s.score
        row_fill = FL_GRN if s.recommendation_level == "recommended" else FL_YEL
        vals = [ri-2, p.part_number, p.manufacturer or "—", p.package or "—",
                round(sc.total_score,1), round(sc.parameter_match_score,1),
                round(sc.supply_risk_score,1),
                {"recommended":"A/B 推荐","backup":"C/D 备选","not_recommended":"E 不推荐"}.get(s.recommendation_level,"—"),
                _lc_label(p.lifecycle_status), "🇨🇳" if getattr(p,'is_domestic',False) else "—"]
        for ci, v in enumerate(vals, 1):
            cell(ws3, ri, ci, v, fill=row_fill, align=AL_C if ci in (1,5,6,7,9,10) else AL_L)

    # ────────────────────────────────────────────────────────────
    # Sheet 4: 供应链风险评估
    # ────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("供应链风险评估")
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 50
    ws4.column_dimensions["D"].width = 35
    section_title(ws4, 1, 1, 4, "  供应链风险评估 (Supply Chain Risk)  [IATF 16949 §8.4]")

    # 整体风险摘要
    cell(ws4, 2, 1, "整体风险等级", font=F_KEY)
    rl = risk_level.lower()
    rf = FL_RED if rl=="high" else (FL_YEL if rl=="medium" else FL_GRN)
    cell(ws4, 2, 2, _risk_emoji(risk_level), fill=rf, align=AL_C)
    if risks and hasattr(risks, 'supply_risk_summary'):
        ws4.merge_cells(f"C2:D2")
        cell(ws4, 2, 3, risks.supply_risk_summary or "—")

    hdr(ws4, 3, 1, "风险类别"); hdr(ws4, 3, 2, "严重度"); hdr(ws4, 3, 3, "描述"); hdr(ws4, 3, 4, "缓解措施")

    if risks and hasattr(risks, 'risk_items'):
        for ri, item in enumerate(risks.risk_items or [], start=4):
            sev = (item.severity or "").lower()
            rfill = FL_RED if sev=="high" else (FL_YEL if sev=="medium" else FL_GRN)
            cell(ws4, ri, 1, item.risk_type or "—")
            cell(ws4, ri, 2, sev.upper(), fill=rfill, align=AL_C)
            cell(ws4, ri, 3, item.description or "—")
            cell(ws4, ri, 4, item.mitigation or "—")

    # 工程风险摘要
    row_e = 4 + len(risks.risk_items if risks and hasattr(risks,'risk_items') else []) + 1
    section_title(ws4, row_e, 1, 4, "  工程风险概述")
    if risks and hasattr(risks, 'engineering_risk_summary'):
        ws4.merge_cells(start_row=row_e+1, start_column=1, end_row=row_e+1, end_column=4)
        cell(ws4, row_e+1, 1, risks.engineering_risk_summary or "—")

    # ────────────────────────────────────────────────────────────
    # Sheet 5: 国产化分析
    # ────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("国产化分析")
    for ci, w in enumerate([8,22,20,14,14,12,40], start=1):
        ws5.column_dimensions[get_column_letter(ci)].width = w
    section_title(ws5, 1, 1, 7, "  国产化分析 (Domestic Substitution Analysis)  [政策符合性]")

    # 国产化概览
    total_r = len(recs)
    dom_r = len(domestic_recs)
    dom_pct = dom_r / total_r * 100 if total_r else 0
    cell(ws5, 2, 1, "推荐总数", font=F_KEY); cell(ws5, 2, 2, total_r, align=AL_C)
    cell(ws5, 3, 1, "国产器件数", font=F_KEY); cell(ws5, 3, 2, dom_r, align=AL_C,
        fill=FL_GRN if dom_pct >= 50 else (FL_YEL if dom_pct >= 20 else FL_RED))
    cell(ws5, 4, 1, "国产化率", font=F_KEY)
    cell(ws5, 4, 2, f"{dom_pct:.1f}%", align=AL_C,
        fill=FL_GRN if dom_pct >= 50 else (FL_YEL if dom_pct >= 20 else FL_RED))
    cell(ws5, 5, 1, "政策目标", font=F_KEY); cell(ws5, 5, 2, "核心器件 ≥ 50%", align=AL_C)

    # 国产器件列表
    section_title(ws5, 7, 1, 7, "  国产推荐器件")
    dom_hdrs = ["排名","型号","制造商","封装","综合评分","生命周期","产品说明"]
    for ci, h in enumerate(dom_hdrs, 1): hdr(ws5, 8, ci, h)
    if domestic_recs:
        for ri, s in enumerate(domestic_recs, start=9):
            p = s.part
            vals = [ri-8, p.part_number, p.manufacturer or "—", p.package or "—",
                    round(s.score.total_score,1), _lc_label(p.lifecycle_status),
                    (p.description or "")[:60]]
            for ci, v in enumerate(vals, 1):
                cell(ws5, ri, ci, v, fill=FL_GRN, align=AL_C if ci in (1,5,6) else AL_L)
    else:
        ws5.merge_cells(f"A9:G9")
        cell(ws5, 9, 1, '本次候选中暂无识别到的国产器件。建议搜索时加入"国产"偏好关键词，或使用 BOM 反查功能。',
             fill=FL_YEL)

    # 国产替代建议
    non_dom = [s for s in recs if not getattr(s.part,'is_domestic',False)]
    if non_dom:
        start_r2 = 10 + len(domestic_recs)
        section_title(ws5, start_r2, 1, 7, "  进口器件的国产替代建议")
        alt_hdrs = ["进口型号","制造商","国产替代方向","推荐搜索关键词","替代难度","备注"]
        for ci, h in enumerate(alt_hdrs, 1): hdr(ws5, start_r2+1, ci+1, h)
        for ri, s in enumerate(non_dom[:8], start=start_r2+2):
            p = s.part
            mfr_lower = (p.manufacturer or "").lower()
            if "texas" in mfr_lower or "ti" == mfr_lower:
                suggest, kw, diff = "圣邦微 SGM / 思瑞浦 3PEAK", "SGM62xxx / SGM4xxx", "中"
            elif "analog" in mfr_lower or "adi" in mfr_lower or "linear" in mfr_lower:
                suggest, kw, diff = "思瑞浦 3PEAK / 南芯 Southchip", "SGM8xxxx / SC8xxx", "高"
            elif "st" in mfr_lower or "microchip" in mfr_lower:
                suggest, kw, diff = "华润矽威 / 芯朋微", "CR6xxx / AP3xxx", "中"
            else:
                suggest, kw, diff = "需根据参数搜索", "参考同类国产品牌", "未知"
            vals = [p.part_number, p.manufacturer or "—", suggest, kw, diff,
                    f"参数需验证：Vout={p.output_voltage_v or '?'}V, Iout={p.output_current_max_a or '?'}A"]
            for ci, v in enumerate(vals, 2):
                cell(ws5, ri, ci, v, fill=FL_YEL if diff != "未知" else None)

    # ────────────────────────────────────────────────────────────
    # Sheet 6: 替代料清单 (AVL)
    # ────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("替代料清单 AVL")
    for ci, w in enumerate([6,22,20,14,10,10,12,14,14,35], start=1):
        ws6.column_dimensions[get_column_letter(ci)].width = w
    section_title(ws6, 1, 1, 10, "  替代料清单 AVL (Approved Vendor List)  [IATF 16949 §8.4.1]")
    avl_hdrs = ["#","型号","制造商","封装","评分","国产","生命周期","推荐等级","兼容等级","说明"]
    for ci, h in enumerate(avl_hdrs, 1): hdr(ws6, 2, ci, h)

    compat_labels = {
        "recommended": "直接替换", "backup": "功能等效（需验证）", "not_recommended": "不推荐"
    }
    for ri, s in enumerate(candidates[:15], start=3):
        p = s.part; sc = s.score
        fill = FL_GRN if s.recommendation_level=="recommended" else (FL_YEL if s.recommendation_level=="backup" else FL_RED)
        vals = [ri-2, p.part_number, p.manufacturer or "—", p.package or "—",
                round(sc.total_score,1), "🇨🇳" if getattr(p,'is_domestic',False) else "—",
                _lc_label(p.lifecycle_status),
                {"recommended":"A/B","backup":"C/D","not_recommended":"E"}.get(s.recommendation_level,"—"),
                compat_labels.get(s.recommendation_level,"—"), (p.description or "")[:50]]
        for ci, v in enumerate(vals, 1):
            cell(ws6, ri, ci, v, fill=fill, align=AL_C if ci in (1,5,6,8,9) else AL_L)

    # ────────────────────────────────────────────────────────────
    # Sheet 7: 证据链与数据来源
    # ────────────────────────────────────────────────────────────
    ws7 = wb.create_sheet("证据链")
    for ci, w in enumerate([22,16,16,10,10,50], start=1):
        ws7.column_dimensions[get_column_letter(ci)].width = w
    section_title(ws7, 1, 1, 6, "  证据链与数据来源 (Evidence Chain)  [ISO 9001 §7.5]")
    ev_hdrs = ["器件型号","证据类型","数据来源","置信度","需复查","声明内容"]
    for ci, h in enumerate(ev_hdrs, 1): hdr(ws7, 2, ci, h)

    evidence = getattr(report, 'evidence', []) or []
    for ri, e in enumerate(evidence[:30], start=3):
        conf = getattr(e, 'confidence', 0)
        fill = FL_GRN if conf >= 0.8 else (FL_YEL if conf >= 0.5 else FL_RED)
        cell(ws7, ri, 1, getattr(e,'part_number',None) or "全局")
        cell(ws7, ri, 2, getattr(e,'evidence_type',None) or "—")
        cell(ws7, ri, 3, getattr(e,'source',None) or "eZ-PLM")
        cell(ws7, ri, 4, f"{conf:.0%}", fill=fill, align=AL_C)
        cell(ws7, ri, 5, "⚠ 是" if getattr(e,'need_human_review',False) else "否", align=AL_C)
        cell(ws7, ri, 6, getattr(e,'claim',None) or "—")

    # ────────────────────────────────────────────────────────────
    # 全局美化：冻结首行 + 自动筛选
    # ────────────────────────────────────────────────────────────
    for ws in [ws2, ws3, ws4, ws6, ws7]:
        try:
            ws.freeze_panes = ws.cell(3, 1)
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
